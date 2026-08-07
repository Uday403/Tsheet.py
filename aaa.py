from __future__ import annotations

import csv
import io
import os
import re
import zipfile
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
MASTER_TEMPLATE = BASE_DIR / "master_template.xlsm"

PRISMA_SHEET = "Prisma Export - Paste as values"
TRAFFIC_SHEET = "Traffic_Doc"
MULTI_SHEET = "Multi-Ad or Creative Rotation"

# AAA standard pmed:
# DPM_ASM_[LOB]_ZZ_[TACTIC]_DSP_[PLATFORM]_[AUDIENCE]_ZZ

AUDIENCE_ALIASES = {
    "LAL": {
        "lal",
        "lookalike",
        "look a like",
        "look-a-like",
        "look alike",
    },
    "PR": {
        "pr",
        "prospecting",
        "prospect",
        "broad prospecting",
    },
    "RT": {
        "rt",
        "retargeting",
        "retarget",
        "remarketing",
    },
}

# Tokens which are useful in the placement but are NOT part of the pmed.
PMED_IGNORE_TOKENS = {
    "MPC",
    "N/A",
    "NA",
    "TPS",
    "SS",
    "ZZ",
}

GENERIC_CREATIVE_WORDS = {
    "aaa",
    "ace",
    "ad",
    "ads",
    "banner",
    "banners",
    "creative",
    "display",
    "dsp",
    "dv360",
    "image",
    "images",
    "native",
    "standard",
    "static",
    "version",
    "ver",
    "new",
}


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _words(value: str) -> list[str]:
    return [
        word
        for word in re.findall(r"[A-Za-z0-9]+", _clean(value).lower())
        if word
    ]


def _read_uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_csv_rows(uploaded_file) -> list[list[str]]:
    text = _decode_csv(_read_uploaded_bytes(uploaded_file))

    try:
        dialect = csv.Sniffer().sniff(
            text[:10000],
            delimiters=",;\t|",
        )
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(text))

    return [list(row) for row in reader]


def _find_prisma_header(
    raw_rows: list[list[str]],
) -> tuple[int, list[str]]:
    for index, row in enumerate(raw_rows):
        headers = [_clean(cell).replace("\n", " ") for cell in row]
        normalized = {_normalize(header) for header in headers}

        if "placementname" in normalized:
            return index, headers

    raise ValueError(
        "The Prisma header row could not be found. "
        "The uploaded file must contain a Placement Name column."
    )


def _get_record_value(
    record: dict[str, str],
    *possible_headers: str,
) -> str:
    normalized_record = {
        _normalize(key): value
        for key, value in record.items()
    }

    for header in possible_headers:
        value = normalized_record.get(_normalize(header))
        if value is not None and _clean(value):
            return _clean(value)

    return ""


def read_prisma_export(
    uploaded_file,
) -> tuple[list[list[str]], list[dict[str, str]]]:
    raw_rows = _read_csv_rows(uploaded_file)
    header_index, headers = _find_prisma_header(raw_rows)

    records: list[dict[str, str]] = []

    for source_index in range(header_index + 1, len(raw_rows)):
        source_row = raw_rows[source_index]
        padded = source_row + [""] * max(0, len(headers) - len(source_row))
        record = dict(zip(headers, padded[:len(headers)]))

        placement = _get_record_value(record, "Placement Name")
        if not placement:
            continue

        row_type = _get_record_value(
            record,
            "Row Type",
            "Type",
            "Package / Placement",
        ).lower()

        if row_type == "package":
            continue

        if placement.lower().startswith("package:"):
            continue

        record["_placement_name"] = placement
        record["_source_excel_row"] = str(source_index + 1)

        record["_site_name"] = _get_record_value(
            record,
            "Media outlet / Supplier name (ad server)",
            "Media outlet / Supplier name (Prisma)",
            "Site Name",
            "Supplier Name",
            "Vendor",
        )

        record["_placement_id"] = _get_record_value(
            record,
            "Ad server ID",
            "Placement ID",
            "DCM Placement ID",
            "Ad Server Placement ID",
        )

        record["_dimension"] = _get_record_value(
            record,
            "Dimensions",
            "Dimension",
            "Creative size",
            "Creative Size",
            "Size",
        )

        record["_start_date"] = _get_record_value(
            record,
            "Flight start date",
            "Start Date",
            "Placement Start Date",
        )

        record["_end_date"] = _get_record_value(
            record,
            "Flight end date",
            "End Date",
            "Placement End Date",
        )

        records.append(record)

    if not records:
        raise ValueError(
            "Placement Name was found, but no placement rows were detected."
        )

    return raw_rows, records


def _extract_dimension(text: str) -> str:
    match = re.search(
        r"(?<!\d)(\d{2,4})\s*[xX×*]\s*(\d{2,4})(?!\d)",
        _clean(text),
    )
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}"


def placement_dimension(record: dict[str, str]) -> str:
    dimension = _extract_dimension(record.get("_dimension", ""))
    if dimension:
        return dimension

    return _extract_dimension(record.get("_placement_name", ""))


def _extract_audience(placement_name: str) -> str:
    placement_normalized = _normalize(placement_name)
    placement_lower = placement_name.lower()

    # Prefer explicit placement taxonomy codes.
    parts = [part.strip() for part in placement_name.split("_")]

    for part in parts:
        upper = part.upper().strip()
        if upper in ("LAL", "PR", "RT"):
            return upper

    for code, aliases in AUDIENCE_ALIASES.items():
        for alias in aliases:
            if alias in placement_lower:
                return code

        if _normalize(code) in placement_normalized:
            return code

    return ""


def _extract_aaa_taxonomy(
    placement_name: str,
    site_name: str = "",
) -> dict[str, str]:
    """
    AAA placement examples:
      DSP_DV360_EC_TRV_MPC_LAL_...
      DSP_DV360_MULTI_TRV_ZZ_PR_...
      DSP_ACW_CA_MEM_ZZ_PR_...
      DSP_ACW_MULTI_MEM_ZZ_PR_...

    We deliberately do not include MPC in pmed.
    """
    parts = [part.strip() for part in placement_name.split("_")]

    platform = ""
    tactic = ""
    lob = ""

    # Most AAA naming follows DSP_[PLATFORM]_[TACTIC]_[LOB]_...
    if parts and parts[0].upper() == "DSP":
        if len(parts) > 1:
            platform = parts[1].upper()
        if len(parts) > 2:
            tactic = parts[2].upper()
        if len(parts) > 3:
            lob = parts[3].upper()

    # Site name is a fallback for platform.
    if not platform:
        site = _clean(site_name).upper()

        platform_aliases = {
            "ACCUWEATHER": "ACW",
            "ACW": "ACW",
            "DV360": "DV360",
            "DISPLAY & VIDEO 360": "DV360",
            "THE TRADE DESK": "TTD",
            "TTD": "TTD",
            "YAHOO": "YAHOO",
        }

        for name, code in platform_aliases.items():
            if name in site:
                platform = code
                break

    audience = _extract_audience(placement_name)

    return {
        "lob": lob,
        "tactic": tactic,
        "platform": platform,
        "audience": audience,
    }


def build_aaa_pmed(
    placement_name: str,
    site_name: str = "",
) -> str:
    taxonomy = _extract_aaa_taxonomy(
        placement_name=placement_name,
        site_name=site_name,
    )

    missing = [
        name
        for name in ("lob", "tactic", "platform", "audience")
        if not taxonomy[name]
    ]

    if missing:
        raise ValueError(
            "Unable to build AAA pmed for placement because these taxonomy "
            f"values were not detected: {', '.join(missing)}. "
            f"Placement: {placement_name}"
        )

    return (
        f"DPM_ASM_{taxonomy['lob']}_ZZ_"
        f"{taxonomy['tactic']}_DSP_"
        f"{taxonomy['platform']}_"
        f"{taxonomy['audience']}_ZZ"
    )


def build_aaa_url(
    base_url: str,
    placement_name: str,
    site_name: str = "",
) -> str:
    """
    Keeps existing query parameters such as Campaigncode and replaces/adds pmed.
    """
    base_url = _clean(base_url)

    if not base_url:
        return ""

    pmed = build_aaa_pmed(
        placement_name=placement_name,
        site_name=site_name,
    )

    parts = urlsplit(base_url)
    query_items = parse_qsl(parts.query, keep_blank_values=True)

    # Remove an existing pmed to avoid duplicates.
    query_items = [
        (key, value)
        for key, value in query_items
        if key.lower() != "pmed"
    ]
    query_items.append(("pmed", pmed))

    query = urlencode(query_items, doseq=True, safe="_")

    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            parts.path,
            query,
            parts.fragment,
        )
    )


def _creative_names_from_uploads(
    creative_files: Iterable,
) -> list[str]:
    names: list[str] = []

    for uploaded_file in creative_files or []:
        file_name = Path(uploaded_file.name).name

        if file_name.lower().endswith(".zip"):
            uploaded_file.seek(0)

            with zipfile.ZipFile(uploaded_file, "r") as archive:
                for member in archive.namelist():
                    if member.endswith("/"):
                        continue

                    name = Path(member).name
                    if name:
                        names.append(name)
        else:
            names.append(file_name)

    # Keep order but remove duplicates.
    return list(dict.fromkeys(name for name in names if name))


def creative_version_key(creative_name: str) -> str:
    """
    Groups the same creative version across dimensions.

    Examples:
      160x600V1 -> V1
      300x250V1 -> V1
      160x600_New Evergreen 7.1 -> New Evergreen 7.1
      300x600_MP Hotel August Banners -> MP Hotel August Banners
    """
    stem = Path(creative_name).stem

    stem = re.sub(
        r"(?<!\d)\d{2,4}\s*[xX×*]\s*\d{2,4}(?!\d)",
        "",
        stem,
        count=1,
    )

    stem = re.sub(r"^[\s_\-]+", "", stem)
    stem = re.sub(r"[\s_\-]+$", "", stem)

    # Common case: 160x600V1 -> V1 after dimension removal.
    return stem or Path(creative_name).stem


def _creative_content_tokens(name: str) -> set[str]:
    stem = Path(name).stem

    # Remove dimensions.
    stem = re.sub(
        r"(?<!\d)\d{2,4}\s*[xX×*]\s*\d{2,4}(?!\d)",
        " ",
        stem,
    )

    tokens = set()

    for token in _words(stem):
        if len(token) < 3:
            continue
        if token in GENERIC_CREATIVE_WORDS:
            continue
        tokens.add(token)

    return tokens


def _single_creative_score(
    creative_name: str,
    placement_name: str,
    required_dimension: str,
) -> float:
    creative_dimension = _extract_dimension(creative_name)
    placement_normalized = _normalize(placement_name)
    creative_stem = Path(creative_name).stem
    score = 0.0

    # For non-1x1 placements, dimension is mandatory.
    if required_dimension and required_dimension.lower() != "1x1":
        if not creative_dimension:
            return -1000
        if creative_dimension.lower() != required_dimension.lower():
            return -1000
        score += 30

    # Native/1x1 placements commonly point to image assets such as 1200x628.
    # In those cases content similarity is more important than 1x1.
    content_tokens = _creative_content_tokens(creative_name)

    for token in content_tokens:
        if _normalize(token) in placement_normalized:
            score += 8

    normalized_stem = _normalize(creative_stem)

    # Strong substring match for names such as TropicalCaribbeanAntilles.
    for segment in re.split(r"[_\-\s]+", creative_stem):
        segment_normalized = _normalize(segment)
        if len(segment_normalized) >= 7 and segment_normalized in placement_normalized:
            score += 20

    # Small bonus for matching audience-related or campaign words.
    overlap = set(_words(creative_stem)) & set(_words(placement_name))
    score += min(len(overlap), 6) * 2

    if normalized_stem and normalized_stem in placement_normalized:
        score += 25

    return score


def match_single_creative(
    creative_names: list[str],
    placement_name: str,
    required_dimension: str,
) -> tuple[str, list[str]]:
    ranked = sorted(
        (
            (
                _single_creative_score(
                    creative_name=name,
                    placement_name=placement_name,
                    required_dimension=required_dimension,
                ),
                name,
            )
            for name in creative_names
        ),
        reverse=True,
    )

    if not ranked or ranked[0][0] <= 0:
        return "", []

    best_score = ranked[0][0]
    tied = [
        name
        for score, name in ranked
        if score == best_score
    ]

    if len(tied) > 1:
        return "", tied

    return ranked[0][1], []


def match_multi_creatives(
    creative_names: list[str],
    required_dimension: str,
) -> list[str]:
    """
    Multi rule confirmed by the user:
    all creatives with the matching placement dimension go under the same Ad.
    """
    if not required_dimension:
        return []

    matches = [
        name
        for name in creative_names
        if _extract_dimension(name).lower() == required_dimension.lower()
    ]

    return matches


def preview_aaa_setup(
    prisma_file,
    creative_files,
    creative_setup: str,
) -> dict:
    _, records = read_prisma_export(prisma_file)
    creative_names = _creative_names_from_uploads(creative_files)

    placement_preview = []
    warnings = []

    for record in records:
        placement_name = record["_placement_name"]
        dimension = placement_dimension(record)

        if creative_setup == "Multiple creatives per ad":
            matches = match_multi_creatives(
                creative_names,
                dimension,
            )
        else:
            creative, tied = match_single_creative(
                creative_names=creative_names,
                placement_name=placement_name,
                required_dimension=dimension,
            )
            matches = [creative] if creative else []

            if tied:
                warnings.append(
                    f"Ambiguous single-creative match for {placement_name}: "
                    + ", ".join(tied)
                )

        placement_preview.append(
            {
                "placement_name": placement_name,
                "dimension": dimension,
                "matches": matches,
            }
        )

    version_groups: dict[str, list[str]] = {}

    for name in creative_names:
        key = creative_version_key(name)
        version_groups.setdefault(key, []).append(name)

    return {
        "placements": placement_preview,
        "creative_names": creative_names,
        "version_groups": version_groups,
        "warnings": warnings,
    }


def _find_traffic_layout(sheet) -> tuple[int, int]:
    """
    Finds Traffic_Doc header row dynamically.
    Returns (header_row, first_data_row).
    """
    for row in range(1, min(sheet.max_row, 30) + 1):
        d_value = _clean(sheet.cell(row=row, column=4).value).lower()
        h_value = _clean(sheet.cell(row=row, column=8).value).lower()

        if "placement name" in d_value and "ad name" in h_value:
            return row, row + 1

    raise ValueError(
        "Unable to locate Traffic_Doc headers "
        "(Placement Name / AD Name)."
    )


def _snapshot_row_format(
    sheet,
    row_number: int,
    max_column: int,
) -> dict:
    snapshot = {}

    for column in range(1, max_column + 1):
        cell = sheet.cell(row=row_number, column=column)

        snapshot[column] = {
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }

    return {
        "cells": snapshot,
        "height": sheet.row_dimensions[row_number].height,
    }


def _apply_row_format(
    sheet,
    row_number: int,
    snapshot: dict,
) -> None:
    for column, style in snapshot["cells"].items():
        cell = sheet.cell(row=row_number, column=column)

        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])

    sheet.row_dimensions[row_number].height = snapshot["height"]


def _clear_sheet_values(
    sheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def _clear_prisma_sheet(sheet) -> None:
    _clear_sheet_values(
        sheet,
        min_row=1,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max(sheet.max_column, 60),
    )


def _paste_prisma_rows(
    sheet,
    raw_rows: list[list[str]],
) -> None:
    for row_number, row_values in enumerate(raw_rows, start=1):
        for column_number, value in enumerate(row_values, start=1):
            sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )


def _clear_existing_merges_below_header(
    sheet,
    first_data_row: int,
) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= first_data_row:
            sheet.unmerge_cells(str(merged_range))


def _to_excel_date_or_text(value: str):
    value = _clean(value)

    if not value:
        return ""

    for fmt in (
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    return value


def _campaign_name_from_prisma(raw_rows: list[list[str]]) -> str:
    for row in raw_rows:
        if not row:
            continue

        first = _clean(row[0]).lower()

        if first in ("campaign name:", "campaign name"):
            if len(row) > 1:
                return _clean(row[1])

    return ""


def _resolve_dates(
    record: dict[str, str],
    override_start_date=None,
    override_end_date=None,
):
    start = override_start_date or _to_excel_date_or_text(
        record.get("_start_date", "")
    )
    end = override_end_date or _to_excel_date_or_text(
        record.get("_end_date", "")
    )
    return start, end


def _version_rotation_for_creative(
    creative_name: str,
    rotation_by_version: dict[str, float],
) -> float | None:
    key = creative_version_key(creative_name)

    value = rotation_by_version.get(key)
    if value is None:
        return None

    # Dashboard uses human percentages, e.g. 25. Convert to Excel 0.25.
    return float(value) / 100.0


def _version_url_for_creative(
    creative_name: str,
    default_base_url: str,
    separate_base_url_by_version: dict[str, str],
) -> str:
    key = creative_version_key(creative_name)
    separate = _clean(separate_base_url_by_version.get(key, ""))

    return separate or default_base_url


def _populate_traffic_sheet(
    workbook,
    records: list[dict[str, str]],
    creative_names: list[str],
    creative_setup: str,
    default_base_url: str,
    override_start_date=None,
    override_end_date=None,
) -> list[str]:
    sheet = workbook[TRAFFIC_SHEET]
    header_row, first_data_row = _find_traffic_layout(sheet)

    style_source_row = first_data_row
    style_snapshot = _snapshot_row_format(
        sheet,
        style_source_row,
        max(sheet.max_column, 24),
    )

    # Remove every old data value.
    _clear_sheet_values(
        sheet,
        min_row=first_data_row,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max(sheet.max_column, 40),
    )

    warnings = []

    for index, record in enumerate(records):
        row = first_data_row + index
        _apply_row_format(sheet, row, style_snapshot)

        placement_name = record["_placement_name"]
        dimension = placement_dimension(record)
        site_name = record["_site_name"]
        start_date, end_date = _resolve_dates(
            record,
            override_start_date,
            override_end_date,
        )

        try:
            final_url = build_aaa_url(
                base_url=default_base_url,
                placement_name=placement_name,
                site_name=site_name,
            )
        except ValueError as exc:
            final_url = ""
            warnings.append(str(exc))

        # Placement Name = Ad Name for AAA.
        ad_name = placement_name

        # Common output.
        sheet.cell(row=row, column=2).value = site_name
        sheet.cell(row=row, column=3).value = record["_placement_id"]
        sheet.cell(row=row, column=4).value = placement_name
        sheet.cell(row=row, column=5).value = dimension
        sheet.cell(row=row, column=8).value = ad_name
        sheet.cell(row=row, column=10).value = "New"
        sheet.cell(row=row, column=12).value = "N"
        sheet.cell(row=row, column=14).value = start_date
        sheet.cell(row=row, column=15).value = end_date
        sheet.cell(row=row, column=16).value = final_url

        if creative_setup == "Multiple creatives per ad":
            matches = match_multi_creatives(
                creative_names,
                dimension,
            )

            if not matches:
                warnings.append(
                    f"No Multi creative matched dimension {dimension}: "
                    f"{placement_name}"
                )

            sheet.cell(row=row, column=11).value = "See Multi Tab"
            sheet.cell(row=row, column=13).value = "SEE MULTI TAB"

        else:
            creative, tied = match_single_creative(
                creative_names=creative_names,
                placement_name=placement_name,
                required_dimension=dimension,
            )

            if tied:
                warnings.append(
                    f"Ambiguous creative match — manual review required: "
                    f"{placement_name} -> {', '.join(tied)}"
                )

            if not creative:
                warnings.append(
                    f"No single creative matched: {placement_name}"
                )

            sheet.cell(row=row, column=11).value = creative
            sheet.cell(row=row, column=13).value = "100%"

            creative_dimension = _extract_dimension(creative)

            if (
                creative
                and dimension
                and dimension.lower() != "1x1"
                and creative_dimension
                and creative_dimension.lower() != dimension.lower()
            ):
                warnings.append(
                    f"Dimension mismatch: placement {dimension}, "
                    f"creative {creative_dimension} — {placement_name}"
                )

    return warnings


def _populate_multi_sheet(
    workbook,
    records: list[dict[str, str]],
    creative_names: list[str],
    default_base_url: str,
    rotation_by_version: dict[str, float],
    separate_base_url_by_version: dict[str, str],
    override_start_date=None,
    override_end_date=None,
) -> list[str]:
    sheet = workbook[MULTI_SHEET]

    first_data_row = 2
    max_column = max(sheet.max_column, 16)

    # Use row 2 as style template.
    style_snapshot = _snapshot_row_format(
        sheet,
        first_data_row,
        max_column,
    )

    # Remove old merges/data.
    _clear_existing_merges_below_header(
        sheet,
        first_data_row,
    )

    _clear_sheet_values(
        sheet,
        min_row=first_data_row,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max_column,
    )

    warnings = []
    output_row = first_data_row

    for record in records:
        placement_name = record["_placement_name"]
        dimension = placement_dimension(record)
        site_name = record["_site_name"]

        matches = match_multi_creatives(
            creative_names,
            dimension,
        )

        if not matches:
            warnings.append(
                f"No creatives matched Multi placement dimension "
                f"{dimension}: {placement_name}"
            )
            continue

        start_date, end_date = _resolve_dates(
            record,
            override_start_date,
            override_end_date,
        )

        block_start = output_row

        for creative_name in matches:
            _apply_row_format(
                sheet,
                output_row,
                style_snapshot,
            )

            rotation = _version_rotation_for_creative(
                creative_name,
                rotation_by_version,
            )

            if rotation is None:
                warnings.append(
                    f"Rotation % missing for creative version "
                    f"'{creative_version_key(creative_name)}'."
                )
                rotation = 0

            base_url = _version_url_for_creative(
                creative_name=creative_name,
                default_base_url=default_base_url,
                separate_base_url_by_version=separate_base_url_by_version,
            )

            try:
                final_url = build_aaa_url(
                    base_url=base_url,
                    placement_name=placement_name,
                    site_name=site_name,
                )
            except ValueError as exc:
                final_url = ""
                warnings.append(str(exc))

            sheet.cell(row=output_row, column=4).value = creative_name
            sheet.cell(row=output_row, column=5).value = "N"
            sheet.cell(row=output_row, column=6).value = rotation
            sheet.cell(row=output_row, column=7).value = start_date
            sheet.cell(row=output_row, column=8).value = end_date
            sheet.cell(row=output_row, column=9).value = final_url

            # Rotation shown as percentage.
            sheet.cell(row=output_row, column=6).number_format = "0%"

            output_row += 1

        block_end = output_row - 1

        # Ad Name merged vertically exactly like AAA examples.
        sheet.cell(row=block_start, column=1).value = placement_name

        if block_end > block_start:
            sheet.merge_cells(
                start_row=block_start,
                start_column=1,
                end_row=block_end,
                end_column=1,
            )

    return warnings


def validate_multi_rotation(
    preview: dict,
    rotation_by_version: dict[str, float],
) -> list[str]:
    """
    Checks that the version rotation applied to each placement totals 100%.
    """
    errors = []

    for placement in preview["placements"]:
        matches = placement["matches"]

        if not matches:
            continue

        total = 0.0
        missing = []

        for creative in matches:
            key = creative_version_key(creative)
            value = rotation_by_version.get(key)

            if value is None:
                missing.append(key)
            else:
                total += float(value)

        if missing:
            errors.append(
                f"{placement['placement_name']}: missing rotation for "
                + ", ".join(sorted(set(missing)))
            )
            continue

        if abs(total - 100.0) > 0.01:
            errors.append(
                f"{placement['placement_name']}: rotation totals "
                f"{total:.2f}% instead of 100%."
            )

    return errors


def generate_aaa_tsheet(
    prisma_file,
    creative_files,
    creative_setup: str,
    default_base_url: str,
    rotation_by_version: dict[str, float] | None = None,
    separate_base_url_by_version: dict[str, str] | None = None,
    override_start_date=None,
    override_end_date=None,
) -> tuple[bytes, list[str]]:
    if not MASTER_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Master template not found: {MASTER_TEMPLATE.name}"
        )

    if creative_setup not in (
        "Single creative per ad",
        "Multiple creatives per ad",
    ):
        raise ValueError("Invalid AAA creative setup type.")

    raw_rows, records = read_prisma_export(prisma_file)
    creative_names = _creative_names_from_uploads(creative_files)

    if not creative_names:
        raise ValueError("No creative files were found.")

    workbook = load_workbook(
        MASTER_TEMPLATE,
        keep_vba=True,
    )

    for required_sheet in (
        PRISMA_SHEET,
        TRAFFIC_SHEET,
        MULTI_SHEET,
    ):
        if required_sheet not in workbook.sheetnames:
            raise KeyError(
                f"Missing worksheet in master template: {required_sheet}"
            )

    # Clear/paste Prisma.
    prisma_sheet = workbook[PRISMA_SHEET]
    _clear_prisma_sheet(prisma_sheet)
    _paste_prisma_rows(prisma_sheet, raw_rows)

    # Campaign-level name if available.
    campaign_name = _campaign_name_from_prisma(raw_rows)

    if campaign_name:
        traffic_sheet = workbook[TRAFFIC_SHEET]
        traffic_sheet["B1"] = campaign_name

    warnings = _populate_traffic_sheet(
        workbook=workbook,
        records=records,
        creative_names=creative_names,
        creative_setup=creative_setup,
        default_base_url=default_base_url,
        override_start_date=override_start_date,
        override_end_date=override_end_date,
    )

    if creative_setup == "Multiple creatives per ad":
        preview = preview_aaa_setup(
            prisma_file=prisma_file,
            creative_files=creative_files,
            creative_setup=creative_setup,
        )

        rotation_by_version = rotation_by_version or {}
        separate_base_url_by_version = (
            separate_base_url_by_version or {}
        )

        rotation_errors = validate_multi_rotation(
            preview=preview,
            rotation_by_version=rotation_by_version,
        )

        if rotation_errors:
            raise ValueError(
                "AAA Multi rotation validation failed:\n"
                + "\n".join(rotation_errors)
            )

        warnings.extend(
            _populate_multi_sheet(
                workbook=workbook,
                records=records,
                creative_names=creative_names,
                default_base_url=default_base_url,
                rotation_by_version=rotation_by_version,
                separate_base_url_by_version=(
                    separate_base_url_by_version
                ),
                override_start_date=override_start_date,
                override_end_date=override_end_date,
            )
        )
    else:
        # Single setup should not retain old Multi-tab campaign data.
        multi_sheet = workbook[MULTI_SHEET]
        _clear_existing_merges_below_header(
            multi_sheet,
            2,
        )
        _clear_sheet_values(
            multi_sheet,
            min_row=2,
            max_row=max(multi_sheet.max_row, 5000),
            min_col=1,
            max_col=max(multi_sheet.max_column, 16),
        )

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue(), warnings
