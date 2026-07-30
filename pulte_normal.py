from __future__ import annotations

import csv
import io
import re
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


BASE_DIR = Path(__file__).resolve().parent
MASTER_TEMPLATE = BASE_DIR / "master_template.xlsm"

PRISMA_SHEET = "Prisma Export - Paste as values"
TRAFFIC_SHEET = "Traffic_Doc"
ROTATION_SHEET = "Multi-Ad or Creative Rotation"

TRAFFIC_HEADER_ROW = 6
TRAFFIC_FIRST_DATA_ROW = 7

# Only these Traffic_Doc columns are manually populated for normal Pulte.
AD_NAME_COLUMN = 8          # H
ACTION_COLUMN = 10          # J
CREATIVE_COLUMN = 11        # K
STUDIO_COLUMN = 12          # L
CLICK_URL_COLUMN = 16       # P


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_prisma_csv(uploaded_file) -> tuple[list[list[str]], list[dict[str, str]]]:
    raw_text = _decode_csv(_uploaded_bytes(uploaded_file))

    try:
        dialect = csv.Sniffer().sniff(
            raw_text[:10000],
            delimiters=",;\t|",
        )
        reader = csv.reader(io.StringIO(raw_text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(raw_text))

    raw_rows = list(reader)

    header_index = None
    placement_header = None

    for row_index, row in enumerate(raw_rows):
        cleaned = [_clean(cell).replace("\n", " ") for cell in row]
        normalized = {
            _normalize(header): header
            for header in cleaned
            if header
        }

        if "placementname" in normalized:
            header_index = row_index
            placement_header = normalized["placementname"]
            break

    if header_index is None or placement_header is None:
        raise ValueError(
            "The Prisma header row could not be found. "
            "The file must contain a Placement Name column."
        )

    headers = [
        _clean(value).replace("\n", " ")
        for value in raw_rows[header_index]
    ]

    records: list[dict[str, str]] = []

    for row in raw_rows[header_index + 1:]:
        padded = row + [""] * max(0, len(headers) - len(row))
        record = dict(zip(headers, padded[:len(headers)]))

        placement_name = _clean(record.get(placement_header))
        if not placement_name:
            continue

        if placement_header != "Placement Name":
            record["Placement Name"] = placement_name

        records.append(record)

    if not records:
        raise ValueError(
            "Placement Name was found, but there are no placement rows below it."
        )

    return raw_rows, records


def _find_placement_name_column(headers: list[str]) -> int:
    for index, header in enumerate(headers, start=1):
        if _normalize(header) == "placementname":
            return index
    raise ValueError("Placement Name column could not be identified.")


def _remove_package_rows(
    raw_rows: list[list[str]],
) -> list[list[str]]:
    """
    Removes package/header rows while preserving the original Prisma layout.

    A valid placement row must have a Placement Name and must not be a package.
    """
    header_index = None

    for index, row in enumerate(raw_rows):
        if any(_normalize(cell) == "placementname" for cell in row):
            header_index = index
            break

    if header_index is None:
        return raw_rows

    headers = [_clean(cell) for cell in raw_rows[header_index]]
    placement_col = _find_placement_name_column(headers) - 1

    output = raw_rows[:header_index + 1]

    for row in raw_rows[header_index + 1:]:
        placement = _clean(row[placement_col]) if placement_col < len(row) else ""
        normalized = placement.lower()

        if not placement:
            continue

        if "package" in normalized and "placement" not in normalized:
            continue

        output.append(row)

    return output


def _clear_prisma_sheet(sheet) -> None:
    max_row = max(sheet.max_row, 5000)
    max_col = max(sheet.max_column, 60)

    for row in sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def paste_prisma_export(
    workbook,
    raw_rows: list[list[str]],
) -> None:
    if PRISMA_SHEET not in workbook.sheetnames:
        raise KeyError(f"Missing worksheet: {PRISMA_SHEET}")

    sheet = workbook[PRISMA_SHEET]
    _clear_prisma_sheet(sheet)

    filtered_rows = _remove_package_rows(raw_rows)

    for row_index, values in enumerate(filtered_rows, start=1):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )


def _find_dimension(value: str) -> str:
    match = re.search(
        r"(?<!\d)(\d{2,4})\s*[xX]\s*(\d{2,4})(?!\d)",
        value,
    )
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}"


def _split_placement(placement_name: str) -> list[str]:
    return [
        _clean(part)
        for part in placement_name.split("_")
    ]


def _parse_placement(placement_name: str) -> dict[str, str]:
    parts = _split_placement(placement_name)

    brand_names = {
        "pulte": "Pulte",
        "delwebb": "Del Webb",
        "centex": "Centex",
        "divosta": "DiVosta",
        "johnwieland": "John Wieland",
        "wieland": "Wieland",
    }

    brand_index = None
    brand = "Pulte"

    for index, part in enumerate(parts):
        normalized = _normalize(part)
        if normalized in brand_names:
            brand_index = index
            brand = brand_names[normalized]
            break

    division = ""
    campaign = ""
    community = ""

    if brand_index is not None:
        if brand_index > 0:
            division = parts[brand_index - 1]

        if brand_index + 1 < len(parts):
            campaign = parts[brand_index + 1]

        if brand_index + 2 < len(parts):
            community = parts[brand_index + 2]

    community_id = ""
    for part in reversed(parts):
        match = re.search(r"(?<!\d)(\d{5,7})(?!\d)", part)
        if match:
            community_id = match.group(1)
            break

    return {
        "division": division,
        "brand": brand,
        "campaign": campaign,
        "community": community,
        "community_id": community_id,
        "dimension": _find_dimension(placement_name),
    }


def build_ad_name(placement_name: str) -> str:
    parsed = _parse_placement(placement_name)

    parts = [
        parsed["division"],
        parsed["brand"],
        parsed["campaign"],
        parsed["community"],
        parsed["dimension"],
    ]

    return "_".join(part for part in parts if part)


def _creative_names(creative_files: Iterable) -> list[str]:
    return [
        Path(file.name).name
        for file in creative_files
        if getattr(file, "name", "")
    ]


def _creative_score(
    creative_name: str,
    placement_name: str,
) -> int:
    parsed = _parse_placement(placement_name)
    normalized_creative = _normalize(creative_name)
    score = 0

    checks = (
        (parsed["dimension"], 15),
        (parsed["community"], 12),
        (parsed["community_id"], 10),
        (parsed["division"], 4),
        (parsed["brand"], 3),
    )

    for value, points in checks:
        normalized_value = _normalize(value)
        if normalized_value and normalized_value in normalized_creative:
            score += points

    community_words = [
        word
        for word in re.split(r"\W+", parsed["community"].lower())
        if len(word) >= 4
    ]

    for word in community_words:
        if word in creative_name.lower():
            score += 3

    return score


def match_creative(
    creative_names: list[str],
    placement_name: str,
) -> str:
    ranked = sorted(
        (
            (_creative_score(name, placement_name), name)
            for name in creative_names
        ),
        reverse=True,
    )

    if not ranked or ranked[0][0] <= 0:
        return ""

    return ranked[0][1]


def _parse_full_urls(url_text: str) -> list[str]:
    urls = []

    for line in url_text.splitlines():
        value = line.strip()
        if value:
            urls.append(value)

    return urls


def match_full_url(
    urls: list[str],
    placement_name: str,
) -> str:
    """
    Normal Pulte receives complete URLs/UTMs from the team.
    This function never creates, edits, or appends a UTM.
    """
    parsed = _parse_placement(placement_name)

    community_id = parsed["community_id"]
    community_words = [
        word.lower()
        for word in re.split(r"\W+", parsed["community"])
        if len(word) >= 4
    ]

    if community_id:
        for url in urls:
            if community_id in url:
                return url

    for url in urls:
        lowered = url.lower()
        if community_words and all(
            word in lowered
            for word in community_words
        ):
            return url

    brand_domain = {
        "Pulte": "pulte.com",
        "Del Webb": "delwebb.com",
        "Centex": "centex.com",
        "DiVosta": "divosta.com",
    }.get(parsed["brand"])

    if brand_domain:
        matching_brand_urls = [
            url for url in urls
            if brand_domain in url.lower()
        ]
        if len(matching_brand_urls) == 1:
            return matching_brand_urls[0]

    if len(urls) == 1:
        return urls[0]

    return ""


def _copy_cell_formula_and_style(
    source,
    target,
) -> None:
    if source.has_style:
        target._style = copy(source._style)

    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)

    if source.data_type == "f" and source.value:
        try:
            target.value = Translator(
                source.value,
                origin=source.coordinate,
            ).translate_formula(target.coordinate)
        except Exception:
            target.value = source.value
    else:
        target.value = source.value


def _prepare_traffic_rows(
    sheet,
    required_rows: int,
) -> None:
    """
    Copies the template formulas and formatting from row 7 down.
    Manual columns are cleared separately.
    """
    last_required_row = (
        TRAFFIC_FIRST_DATA_ROW + required_rows - 1
    )

    source_row = TRAFFIC_FIRST_DATA_ROW

    for target_row in range(
        TRAFFIC_FIRST_DATA_ROW,
        last_required_row + 1,
    ):
        if target_row != source_row:
            for column in range(1, 24):
                _copy_cell_formula_and_style(
                    sheet.cell(source_row, column),
                    sheet.cell(target_row, column),
                )

            sheet.row_dimensions[target_row].height = (
                sheet.row_dimensions[source_row].height
            )


def _clear_old_manual_traffic_data(sheet) -> None:
    """
    Clears only editable/manual fields.

    Formula columns such as Site Name, Placement ID, Placement Name,
    Dimensions, Trafficking Notes, Rotation and Dates are preserved.
    """
    max_row = max(sheet.max_row, 5000)

    manual_columns = [
        AD_NAME_COLUMN,
        ACTION_COLUMN,
        CREATIVE_COLUMN,
        STUDIO_COLUMN,
        16, 17, 18, 19, 20, 21, 22, 23,  # P:W clickTag URLs
    ]

    for row in range(TRAFFIC_FIRST_DATA_ROW, max_row + 1):
        for column in manual_columns:
            sheet.cell(row=row, column=column).value = None


def _clear_old_rotation_data(workbook) -> None:
    if ROTATION_SHEET not in workbook.sheetnames:
        return

    sheet = workbook[ROTATION_SHEET]
    max_row = max(sheet.max_row, 5000)
    max_col = max(sheet.max_column, 8)

    for row in sheet.iter_rows(
        min_row=2,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def populate_normal_pulte_sheet(
    workbook,
    records: list[dict[str, str]],
    creative_files: Iterable,
    complete_urls_text: str,
) -> list[str]:
    if TRAFFIC_SHEET not in workbook.sheetnames:
        raise KeyError(f"Missing worksheet: {TRAFFIC_SHEET}")

    sheet = workbook[TRAFFIC_SHEET]

    _clear_old_manual_traffic_data(sheet)
    _prepare_traffic_rows(sheet, len(records))
    _clear_old_rotation_data(workbook)

    creatives = _creative_names(creative_files)
    urls = _parse_full_urls(complete_urls_text)
    warnings: list[str] = []

    for index, record in enumerate(records):
        row = TRAFFIC_FIRST_DATA_ROW + index
        placement_name = _clean(record.get("Placement Name"))

        ad_name = build_ad_name(placement_name)
        creative_name = match_creative(
            creatives,
            placement_name,
        )
        full_url = match_full_url(
            urls,
            placement_name,
        )

        sheet.cell(row=row, column=AD_NAME_COLUMN).value = ad_name
        sheet.cell(row=row, column=ACTION_COLUMN).value = "New"
        sheet.cell(row=row, column=CREATIVE_COLUMN).value = creative_name
        sheet.cell(row=row, column=STUDIO_COLUMN).value = "N"
        sheet.cell(row=row, column=CLICK_URL_COLUMN).value = full_url

        if not creative_name:
            warnings.append(
                f"No creative matched: {placement_name}"
            )

        expected_dimension = _find_dimension(placement_name)
        creative_dimension = _find_dimension(creative_name)

        if (
            creative_name
            and expected_dimension
            and creative_dimension
            and expected_dimension.lower()
            != creative_dimension.lower()
        ):
            warnings.append(
                "Dimension mismatch: "
                f"{expected_dimension} placement vs "
                f"{creative_dimension} creative — {placement_name}"
            )

        if not full_url:
            warnings.append(
                f"No complete URL/UTM matched: {placement_name}"
            )

    # Remove stale manual values beyond the current number of placements.
    current_last_row = TRAFFIC_FIRST_DATA_ROW + len(records) - 1

    for row in range(current_last_row + 1, max(sheet.max_row, 5000) + 1):
        for column in (
            AD_NAME_COLUMN,
            ACTION_COLUMN,
            CREATIVE_COLUMN,
            STUDIO_COLUMN,
            16, 17, 18, 19, 20, 21, 22, 23,
        ):
            sheet.cell(row=row, column=column).value = None

    return warnings


def generate_normal_pulte_tsheet(
    prisma_file,
    creative_files,
    complete_urls_text: str,
) -> tuple[bytes, list[str]]:
    if not MASTER_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Master template not found: {MASTER_TEMPLATE.name}"
        )

    raw_rows, records = read_prisma_csv(prisma_file)

    workbook = load_workbook(
        MASTER_TEMPLATE,
        keep_vba=True,
    )

    paste_prisma_export(
        workbook,
        raw_rows,
    )

    warnings = populate_normal_pulte_sheet(
        workbook=workbook,
        records=records,
        creative_files=creative_files,
        complete_urls_text=complete_urls_text,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue(), warnings
