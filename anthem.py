from __future__ import annotations

import csv
import io
import re
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


BASE_DIR = Path(__file__).resolve().parent
MASTER_TEMPLATE = BASE_DIR / "master_template.xlsm"

PRISMA_SHEET = "Prisma Export - Paste as values"
TRAFFIC_SHEET = "Traffic_Doc"
MULTI_SHEET = "Multi-Ad or Creative Rotation"

SEE_MULTI = "SEE MULTI-CREATIVE TAB"
ROTATION_MULTI = "SEE MULTI TAB"

# Anthem / Elevance ad-name pattern confirmed from the supplied MDCD workbook:
# ELV_[LOB]_[State]_[Language]_[Channel]_[Size/Duration]
#
# Examples:
# ELV_MDCD_Florida_EN_Display_300x250
# ELV_MDCD_Florida_SP_CTV_:15xNAN
# ELV_MDCD_Florida_EN_Video_:30xNAN

STATE_NAMES = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AZ": "Arizona",
    "AR": "Arkansas",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
    "DC": "District of Columbia",
}

LOB_ALIASES = {
    "medicaid": "MDCD",
    "mdcd": "MDCD",
    "medicare": "MDCR",
    "mdcr": "MDCR",
    "csbd": "CSBD",
}

CREATIVE_LANGUAGE_MARKERS = {
    # Confirmed in the supplied Anthem creative packages.
    "EN": ("FLCENSHP",),
    "SP": ("FLCSPSHP",),
}


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _read_uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def _read_csv_rows(uploaded_file) -> list[list[str]]:
    text = _decode_csv(_read_uploaded_bytes(uploaded_file))

    try:
        dialect = csv.Sniffer().sniff(
            text[:10000],
            delimiters=",;\t|",
        )
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(text))

    return [list(row) for row in reader]


def _find_prisma_header(
    raw_rows: list[list[str]],
) -> tuple[int, list[str]]:
    for index, row in enumerate(raw_rows):
        headers = [_clean(cell).replace("\n", " ") for cell in row]
        normalized = {_normalize(header) for header in headers}

        if "placementname" in normalized:
            return index, headers

    raise ValueError(
        "The Prisma header row could not be found. "
        "The file must contain a Placement Name column."
    )


def _get_record_value(
    record: dict[str, str],
    *possible_headers: str,
) -> str:
    normalized_record = {
        _normalize(key): value
        for key, value in record.items()
    }

    for header in possible_headers:
        value = normalized_record.get(_normalize(header))
        if value is not None and _clean(value):
            return _clean(value)

    return ""


def read_prisma_export(
    uploaded_file,
) -> tuple[list[list[str]], list[dict[str, str]]]:
    raw_rows = _read_csv_rows(uploaded_file)
    header_index, headers = _find_prisma_header(raw_rows)

    records: list[dict[str, str]] = []

    for source_index in range(header_index + 1, len(raw_rows)):
        source_row = raw_rows[source_index]
        padded = source_row + [""] * max(
            0,
            len(headers) - len(source_row),
        )
        record = dict(zip(headers, padded[:len(headers)]))

        placement_name = _get_record_value(
            record,
            "Placement Name",
        )

        if not placement_name:
            # Package rows in the Anthem Prisma export have blank Placement Name.
            continue

        record["_placement_name"] = placement_name
        record["_source_excel_row"] = str(source_index + 1)

        record["_site_name"] = _get_record_value(
            record,
            "Media outlet / Supplier name (ad server)",
            "Media outlet / Supplier name (Prisma)",
            "Site Name",
            "Supplier Name",
            "Vendor",
        )

        record["_placement_id"] = _get_record_value(
            record,
            "Ad server ID",
            "Placement ID",
            "DCM Placement ID",
            "Ad Server Placement ID",
        )

        record["_dimension"] = _get_record_value(
            record,
            "Ad size",
            "Tag size",
            "Dimensions",
            "Dimension",
            "Creative Size",
            "Size",
        )

        record["_placement_type"] = _get_record_value(
            record,
            "Placement type",
            "Placement Type",
            "Type",
        )

        record["_product"] = _get_record_value(
            record,
            "Product",
            "Product Name",
        )

        record["_start_date"] = _get_record_value(
            record,
            "Flight start date",
            "Start Date",
            "Placement Start Date",
        )

        record["_end_date"] = _get_record_value(
            record,
            "Flight end date",
            "End Date",
            "Placement End Date",
        )

        records.append(record)

    if not records:
        raise ValueError(
            "No Anthem placement rows were detected in the Prisma export."
        )

    return raw_rows, records


def _extract_dimension(text: str) -> str:
    match = re.search(
        r"(?<!\d)(\d{2,4})\s*[xX×*]\s*(\d{2,4})(?!\d)",
        _clean(text),
    )
    if not match:
        return ""

    return f"{match.group(1)}x{match.group(2)}"


def placement_dimension(record: dict[str, str]) -> str:
    dimension = _extract_dimension(
        record.get("_dimension", "")
    )
    if dimension:
        return dimension

    return _extract_dimension(
        record.get("_placement_name", "")
    )


def detect_language(placement_name: str) -> str:
    # Anthem language markers can sit inside a taxonomy segment such as
    # "CRM Predictive A19-64 EN", so underscore splitting alone is not enough.
    if re.search(
        r"(?<![A-Za-z0-9])SP(?![A-Za-z0-9])",
        placement_name,
        flags=re.IGNORECASE,
    ):
        return "SP"

    if re.search(
        r"(?<![A-Za-z0-9])EN(?![A-Za-z0-9])",
        placement_name,
        flags=re.IGNORECASE,
    ):
        return "EN"

    lowered = placement_name.lower()

    if "spanish" in lowered:
        return "SP"

    if "english" in lowered:
        return "EN"

    return ""


def detect_channel(record: dict[str, str]) -> str:
    placement_name = record["_placement_name"]
    placement_type = record.get("_placement_type", "")

    combined = f"{placement_type} {placement_name}".lower()

    if "ctv" in combined:
        return "CTV"

    if (
        "digital radio" in combined
        or "digital audio" in combined
        or "audio" in combined
        or "drad" in combined
    ):
        return "Audio"

    if (
        "video" in combined
        or "instream" in combined
        or "olv" in combined
    ):
        return "Video"

    if (
        "display" in combined
        or "banner" in combined
    ):
        return "Display"

    return ""


def detect_video_duration(placement_name: str) -> str:
    # Anthem placement examples contain :15xNAN / :30xNAN.
    match = re.search(
        r":\s*(6|15|30|60|90)\s*x",
        placement_name,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1)

    # Additional safe fallbacks.
    match = re.search(
        r"(?<!\d)(6|15|30|60|90)\s*s(?:ec(?:ond)?s?)?(?!\d)",
        placement_name,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1)

    return ""


def detect_lob(
    record: dict[str, str],
    campaign_name: str = "",
) -> str:
    sources = [
        record.get("_product", ""),
        record.get("_placement_name", ""),
        campaign_name,
    ]

    for source in sources:
        lowered = source.lower()

        # Prefer explicit abbreviations.
        for code in ("MDCD", "MDCR", "CSBD"):
            if re.search(
                rf"(?<![A-Za-z0-9]){code}(?![A-Za-z0-9])",
                source,
                flags=re.IGNORECASE,
            ):
                return code

        for alias, code in LOB_ALIASES.items():
            if alias in lowered:
                return code

    return ""


def detect_state_code(
    record: dict[str, str],
    campaign_name: str = "",
) -> str:
    placement_name = record["_placement_name"]

    # In Anthem taxonomy, state generally appears near the tail:
    # ..._FL_MDCD_...
    parts = [
        part.strip().upper()
        for part in placement_name.split("_")
    ]

    for part in reversed(parts):
        if part in STATE_NAMES:
            return part

    # Campaign fallback such as ANTH_MDCD_FL ...
    campaign_parts = re.split(
        r"[_\s\-]+",
        campaign_name.upper(),
    )

    for part in campaign_parts:
        if part in STATE_NAMES:
            return part

    return ""


def build_anthem_ad_name(
    record: dict[str, str],
    campaign_name: str = "",
) -> str:
    placement_name = record["_placement_name"]

    lob = detect_lob(record, campaign_name)
    state_code = detect_state_code(record, campaign_name)
    language = detect_language(placement_name)
    channel = detect_channel(record)

    missing = []

    if not lob:
        missing.append("LOB")

    if not state_code:
        missing.append("State")

    if not language:
        missing.append("Language")

    if not channel:
        missing.append("Channel")

    if missing:
        raise ValueError(
            "Unable to build Anthem Ad Name. Missing "
            f"{', '.join(missing)} for placement: {placement_name}"
        )

    state_name = STATE_NAMES[state_code]

    if channel == "Display":
        dimension = placement_dimension(record)

        if not dimension:
            raise ValueError(
                "Unable to detect display dimension for placement: "
                f"{placement_name}"
            )

        suffix = dimension

    else:
        duration = detect_video_duration(
            placement_name
        )

        if not duration:
            raise ValueError(
                "Unable to detect video duration for placement: "
                f"{placement_name}"
            )

        # Preserve the format used by the supplied Anthem workbook.
        suffix = f":{duration}xNAN"

    return (
        f"ELV_{lob}_{state_name}_{language}_"
        f"{channel}_{suffix}"
    )


def _campaign_name_from_prisma(
    raw_rows: list[list[str]],
) -> str:
    for row in raw_rows:
        if not row:
            continue

        first = _clean(row[0]).lower()

        if first in ("campaign name:", "campaign name"):
            if len(row) > 1:
                return _clean(row[1])

    return ""


def _creative_names_from_uploads(
    creative_files: Iterable,
) -> list[str]:
    """
    Supports:
    - individual creative files
    - ZIPs containing creative files
    - ZIPs containing other ZIP display assets

    For Anthem Display, the inner ZIP filename itself is the trafficking
    creative filename, so nested .zip names are retained rather than opened.
    """
    names: list[str] = []

    for uploaded_file in creative_files or []:
        file_name = Path(uploaded_file.name).name

        if file_name.lower().endswith(".zip"):
            uploaded_file.seek(0)

            with zipfile.ZipFile(uploaded_file, "r") as archive:
                for member in archive.namelist():
                    if member.endswith("/"):
                        continue

                    member_name = Path(member).name
                    if not member_name:
                        continue

                    # Anthem supplied display creatives are themselves ZIPs.
                    # Video creatives are MP4.
                    if member_name.lower().endswith(
                        (".zip", ".mp4", ".jpg", ".jpeg",
                         ".png", ".gif", ".webp")
                    ):
                        names.append(member_name)

        else:
            names.append(file_name)

    # Keep upload/archive order while removing duplicates.
    return list(dict.fromkeys(
        name for name in names if name
    ))


def creative_language(
    creative_name: str,
) -> str:
    upper = creative_name.upper()

    # Explicit campaign markers confirmed from supplied files.
    if "FLCENSHP" in upper:
        return "EN"

    if "FLCSPSHP" in upper:
        return "SP"

    # Generic fallback for future state campaigns where filenames
    # explicitly contain language.
    tokens = re.split(r"[^A-Z0-9]+", upper)

    if "EN" in tokens or "ENGLISH" in tokens:
        return "EN"

    if "SP" in tokens or "SPANISH" in tokens:
        return "SP"

    return ""


def creative_duration(
    creative_name: str,
) -> str:
    stem = Path(creative_name).stem

    # Example:
    # "... Back to School 15 16x9 VD ..."
    match = re.search(
        r"(?<!\d)(6|15|30|60|90)\s+"
        r"(?:16\s*[xX]\s*9|9\s*[xX]\s*16)(?!\d)",
        stem,
        flags=re.IGNORECASE,
    )

    if match:
        return match.group(1)

    match = re.search(
        r"(?<!\d)(6|15|30|60|90)\s*s(?:ec)?(?!\d)",
        stem,
        flags=re.IGNORECASE,
    )

    return match.group(1) if match else ""


def creative_aspect_ratio(
    creative_name: str,
) -> str:
    match = re.search(
        r"(?<!\d)(16|9)\s*[xX]\s*(9|16)(?!\d)",
        creative_name,
    )

    if not match:
        return ""

    return f"{match.group(1)}x{match.group(2)}"


def match_anthem_creatives(
    record: dict[str, str],
    creative_names: list[str],
) -> list[str]:
    placement_name = record["_placement_name"]
    language = detect_language(placement_name)
    channel = detect_channel(record)

    if not language or not channel:
        return []

    # DISPLAY
    if channel == "Display":
        required_dimension = placement_dimension(record)

        if not required_dimension:
            return []

        matches = []

        for name in creative_names:
            if creative_language(name) != language:
                continue

            if _extract_dimension(name).lower() != required_dimension.lower():
                continue

            if not name.lower().endswith(
                (".zip", ".jpg", ".jpeg", ".png", ".gif", ".webp")
            ):
                continue

            matches.append(name)

        return matches

    required_duration = detect_video_duration(
        placement_name
    )

    if not required_duration:
        return []

    # AUDIO
    # Never allow OLV/CTV MP4 creatives to be adopted by Audio placements.
    # Audio placements may only match actual audio-file formats.
    if channel == "Audio":
        audio_extensions = (
            ".mp3",
            ".wav",
            ".m4a",
            ".aac",
            ".ogg",
        )

        matches = []

        for name in creative_names:
            if not name.lower().endswith(audio_extensions):
                continue

            if creative_language(name) != language:
                continue

            if creative_duration(name) != required_duration:
                continue

            matches.append(name)

        return matches

    # VIDEO / CTV
    # Only actual video files are eligible.
    if channel in ("Video", "CTV"):
        matches = []

        for name in creative_names:
            if not name.lower().endswith(".mp4"):
                continue

            if creative_language(name) != language:
                continue

            if creative_duration(name) != required_duration:
                continue

            # Existing Anthem OLV/CTV trafficking uses 16x9 assets.
            aspect = creative_aspect_ratio(name)

            if aspect and aspect != "16x9":
                continue

            matches.append(name)

        return matches

    return []

def _infer_url_language(url: str) -> str:
    """
    Anthem rule:
      /es/ in URL -> Spanish
      otherwise   -> English
    """
    return "SP" if "/es/" in url.lower() else "EN"


def _infer_url_channel(url: str) -> str:
    """
    Anthem CMP rules:
      DIS  -> Display
      OLV  -> Video
      CTV  -> CTV
      DRAD -> Audio
    """
    upper = url.upper()

    if re.search(r"(?:[?&]CMP=|[-_])DIS-", upper):
        return "Display"

    if "BRC-OLV-" in upper or "CMP=OLV-" in upper:
        return "Video"

    if "BRC-CTV-" in upper or "CMP=CTV-" in upper:
        return "CTV"

    if "BRC-DRAD-" in upper or "CMP=DRAD-" in upper:
        return "Audio"

    return ""


def parse_url_mapping(
    url_mapping_text: str = "",
    url_mapping: dict | None = None,
) -> dict[tuple[str, str], str]:
    """
    The user may simply paste Anthem URLs, one per line.

    Mapping is automatic:
      /es/ -> SP, otherwise EN
      DIS  -> Display
      OLV  -> Video
      CTV  -> CTV
      DRAD -> Audio

    This parser deliberately extracts URLs from anywhere in the pasted
    text, so leading bullets, spaces, numbering, or Excel paste artifacts
    do not prevent mapping.
    """
    result: dict[tuple[str, str], str] = {}

    # Keep backward compatibility with a supplied dict.
    if url_mapping:
        for key, value in url_mapping.items():
            if not (isinstance(key, tuple) and len(key) == 2):
                continue

            language = str(key[0]).upper().strip()
            channel_raw = str(key[1]).strip().upper()

            channel_lookup = {
                "DISPLAY": "Display",
                "DIS": "Display",
                "VIDEO": "Video",
                "OLV": "Video",
                "CTV": "CTV",
                "AUDIO": "Audio",
                "DRAD": "Audio",
            }

            channel = channel_lookup.get(channel_raw, "")
            cleaned_url = _clean(value)

            if (
                language in ("EN", "SP")
                and channel
                and cleaned_url
            ):
                result[(language, channel)] = cleaned_url

    pasted_text = _clean(url_mapping_text)

    # Extract every HTTP/HTTPS URL, regardless of line formatting.
    urls = re.findall(
        r'https?://[^\s<>"\']+',
        pasted_text,
        flags=re.IGNORECASE,
    )

    for url in urls:
        # Remove punctuation that may be copied after a URL.
        url = url.rstrip(".,);]")

        language = _infer_url_language(url)
        channel = _infer_url_channel(url)

        if channel:
            result[(language, channel)] = url

    return result

def _resolve_url(
    record: dict[str, str],
    url_map: dict[tuple[str, str], str],
) -> str:
    language = detect_language(
        record["_placement_name"]
    )
    channel = detect_channel(record)

    return _clean(
        url_map.get((language, channel), "")
    )


def _to_excel_date_or_text(value):
    """
    Normalize every supported date source to a real Python datetime.

    Handles:
    - existing datetime/date values
    - Excel serial dates such as 46273 / 46387
    - numeric serials stored as text
    - mm/dd/yyyy
    - mm-dd-yyyy
    - dd-mm-yyyy
    - dd/mm/yyyy
    - yyyy-mm-dd
    - yyyy/mm/dd
    - yyyy-mm-dd HH:MM:SS
    """
    if value is None:
        return ""

    # Already a real date/datetime.
    if isinstance(value, datetime):
        return value

    # Excel serial supplied as int/float.
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return value

    value = _clean(value)

    if not value:
        return ""

    # Excel serial supplied as text, e.g. "46273" or "46273.0".
    if re.fullmatch(r"\d+(?:\.0+)?", value):
        try:
            serial = float(value)
            # Normal modern Excel dates are safely in this range.
            if 1 <= serial <= 100000:
                return from_excel(serial)
        except Exception:
            pass

    for fmt in (
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    return value

def _resolve_dates(
    record: dict[str, str],
    override_start_date=None,
    override_end_date=None,
):
    start = (
        override_start_date
        if override_start_date is not None
        else _to_excel_date_or_text(
            record.get("_start_date", "")
        )
    )

    end = (
        override_end_date
        if override_end_date is not None
        else _to_excel_date_or_text(
            record.get("_end_date", "")
        )
    )

    return start, end


def preview_anthem_setup(
    prisma_file,
    creative_files,
    url_mapping_text: str = "",
    url_mapping: dict | None = None,
) -> dict:
    raw_rows, records = read_prisma_export(
        prisma_file
    )
    campaign_name = _campaign_name_from_prisma(
        raw_rows
    )

    creative_names = _creative_names_from_uploads(
        creative_files
    )
    url_map = parse_url_mapping(
        url_mapping_text,
        url_mapping,
    )

    rows = []
    warnings = []
    unique_ads: dict[str, dict] = {}

    for record in records:
        placement_name = record["_placement_name"]

        try:
            ad_name = build_anthem_ad_name(
                record,
                campaign_name,
            )
        except ValueError as exc:
            ad_name = ""
            warnings.append(str(exc))

        matches = match_anthem_creatives(
            record,
            creative_names,
        )

        final_url = _resolve_url(
            record,
            url_map,
        )

        if not matches:
            warnings.append(
                "No Anthem creative matched: "
                f"{placement_name}"
            )

        if not final_url:
            language = detect_language(
                placement_name
            )
            channel = detect_channel(record)
            warnings.append(
                f"No URL mapping for {language} {channel}: "
                f"{placement_name}"
            )

        rows.append(
            {
                "placement_name": placement_name,
                "ad_name": ad_name,
                "language": detect_language(
                    placement_name
                ),
                "channel": detect_channel(record),
                "dimension": placement_dimension(
                    record
                ),
                "duration": detect_video_duration(
                    placement_name
                ),
                "matches": matches,
                "creative_destination": (
                    "Multi"
                    if len(matches) >= 2
                    else "Traffic_Doc"
                    if len(matches) == 1
                    else "Unmatched"
                ),
                "url": final_url,
            }
        )

        if ad_name:
            unique_ads.setdefault(
                ad_name,
                {
                    "ad_name": ad_name,
                    "matches": matches,
                    "url": final_url,
                },
            )

    return {
        "campaign_name": campaign_name,
        "placements": rows,
        "unique_ads": list(
            unique_ads.values()
        ),
        "creative_names": creative_names,
        "warnings": list(dict.fromkeys(warnings)),
    }


def _find_traffic_layout(
    sheet,
) -> tuple[int, int]:
    for row in range(
        1,
        min(sheet.max_row, 30) + 1,
    ):
        d_value = _clean(
            sheet.cell(
                row=row,
                column=4,
            ).value
        ).lower()

        h_value = _clean(
            sheet.cell(
                row=row,
                column=8,
            ).value
        ).lower()

        if (
            "placement name" in d_value
            and "ad name" in h_value
        ):
            return row, row + 1

    raise ValueError(
        "Unable to locate Traffic_Doc headers."
    )


def _snapshot_row_format(
    sheet,
    row_number: int,
    max_column: int,
) -> dict:
    snapshot = {}

    for column in range(
        1,
        max_column + 1,
    ):
        cell = sheet.cell(
            row=row_number,
            column=column,
        )

        snapshot[column] = {
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }

    return {
        "cells": snapshot,
        "height": sheet.row_dimensions[
            row_number
        ].height,
    }


def _apply_row_format(
    sheet,
    row_number: int,
    snapshot: dict,
) -> None:
    for column, style in snapshot[
        "cells"
    ].items():
        cell = sheet.cell(
            row=row_number,
            column=column,
        )

        cell._style = copy(style["style"])
        cell.number_format = style[
            "number_format"
        ]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(
            style["alignment"]
        )
        cell.protection = copy(
            style["protection"]
        )

    sheet.row_dimensions[
        row_number
    ].height = snapshot["height"]


def _clear_values(
    sheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    # Only clear the existing used range; do not force 5,000 rows.
    if max_row < min_row:
        return

    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def _paste_prisma_rows(
    sheet,
    raw_rows: list[list[str]],
) -> None:
    for row_number, row_values in enumerate(
        raw_rows,
        start=1,
    ):
        for column_number, value in enumerate(
            row_values,
            start=1,
        ):
            sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )


def _populate_traffic_sheet(
    workbook,
    records: list[dict[str, str]],
    creative_names: list[str],
    campaign_name: str,
    url_map: dict[tuple[str, str], str],
    override_start_date=None,
    override_end_date=None,
) -> list[str]:
    sheet = workbook[TRAFFIC_SHEET]

    # Anthem Traffic_Doc campaign header.
    # B2 must reflect the campaign name read from the Prisma export.
    sheet["B2"] = campaign_name

    _, first_data_row = _find_traffic_layout(
        sheet
    )

    max_column = max(
        24,
        min(sheet.max_column, 40),
    )

    style_snapshot = _snapshot_row_format(
        sheet,
        first_data_row,
        max_column,
    )

    old_max_row = sheet.max_row

    _clear_values(
        sheet,
        min_row=first_data_row,
        max_row=old_max_row,
        min_col=1,
        max_col=max_column,
    )

    warnings = []

    for index, record in enumerate(records):
        row = first_data_row + index

        _apply_row_format(
            sheet,
            row,
            style_snapshot,
        )

        placement_name = record[
            "_placement_name"
        ]
        channel = detect_channel(record)

        try:
            ad_name = build_anthem_ad_name(
                record,
                campaign_name,
            )
        except ValueError as exc:
            ad_name = ""
            warnings.append(str(exc))

        matches = match_anthem_creatives(
            record,
            creative_names,
        )

        start_date, end_date = _resolve_dates(
            record,
            override_start_date,
            override_end_date,
        )

        final_url = _resolve_url(
            record,
            url_map,
        )

        # Shared Traffic_Doc fields.
        sheet.cell(
            row=row,
            column=1,
        ).value = "DoubleVerify"

        sheet.cell(
            row=row,
            column=2,
        ).value = record["_site_name"]

        sheet.cell(
            row=row,
            column=3,
        ).value = record["_placement_id"]

        sheet.cell(
            row=row,
            column=4,
        ).value = placement_name

        if channel == "Display":
            sheet.cell(
                row=row,
                column=5,
            ).value = placement_dimension(record)

            sheet.cell(
                row=row,
                column=6,
            ).value = ""

            sheet.cell(
                row=row,
                column=7,
            ).value = ""

        else:
            sheet.cell(
                row=row,
                column=5,
            ).value = "0x0"

            sheet.cell(
                row=row,
                column=6,
            ).value = detect_video_duration(
                placement_name
            )

            # Supplied template defaults video rows to Vpaid unless
            # platform-specific logic overrides it. Yahoo sample uses Vpaid.
            sheet.cell(
                row=row,
                column=7,
            ).value = "Vpaid"

        sheet.cell(
            row=row,
            column=8,
        ).value = ad_name

        sheet.cell(
            row=row,
            column=10,
        ).value = "New"

        sheet.cell(
            row=row,
            column=12,
        ).value = "N"

        start_cell = sheet.cell(
            row=row,
            column=14,
        )
        start_cell.value = start_date
        start_cell.number_format = "dd-mm-yyyy"

        end_cell = sheet.cell(
            row=row,
            column=15,
        )
        end_cell.value = end_date
        end_cell.number_format = "dd-mm-yyyy"

        if len(matches) == 0:
            sheet.cell(
                row=row,
                column=11,
            ).value = ""

            sheet.cell(
                row=row,
                column=13,
            ).value = ""

            sheet.cell(
                row=row,
                column=16,
            ).value = final_url

            warnings.append(
                "No creative matched Traffic_Doc placement: "
                f"{placement_name}"
            )

        elif len(matches) == 1:
            # One creative = actual creative goes directly into Traffic_Doc.
            sheet.cell(
                row=row,
                column=11,
            ).value = matches[0]

            sheet.cell(
                row=row,
                column=13,
            ).value = "100%"

            sheet.cell(
                row=row,
                column=16,
            ).value = final_url

        else:
            # Two or more creatives = Traffic_Doc points to Multi tab.
            sheet.cell(
                row=row,
                column=11,
            ).value = SEE_MULTI

            sheet.cell(
                row=row,
                column=13,
            ).value = ROTATION_MULTI

            sheet.cell(
                row=row,
                column=16,
            ).value = SEE_MULTI

        if not final_url:
            warnings.append(
                "URL missing for placement: "
                f"{placement_name}"
            )

    return warnings


def _find_header_column(
    sheet,
    *header_candidates: str,
) -> int:
    """
    Find a column by the actual header text in row 1.
    This avoids column-shift problems when account templates differ.
    """
    normalized_candidates = {
        _normalize(candidate)
        for candidate in header_candidates
    }

    for column in range(1, sheet.max_column + 1):
        header = _normalize(
            sheet.cell(
                row=1,
                column=column,
            ).value
        )

        if not header:
            continue

        if header in normalized_candidates:
            return column

    raise ValueError(
        "Could not find Multi-tab column for: "
        + " / ".join(header_candidates)
    )


def _populate_multi_sheet(
    workbook,
    records: list[dict[str, str]],
    creative_names: list[str],
    campaign_name: str,
    url_map: dict[tuple[str, str], str],
    override_start_date=None,
    override_end_date=None,
) -> list[str]:
    sheet = workbook[MULTI_SHEET]

    first_data_row = 2
    max_column = max(
        sheet.max_column,
        16,
    )

    # IMPORTANT:
    # Do not hardcode C/D/E/etc. Use the headers in master_template.xlsm.
    # Current shared master has:
    # A = AD Name
    # B = Trafficking Notes
    # C = Creative File Name
    # D = Studio Creative? (Y/N)
    # E = Rotation %
    # F = Start Date
    # G = End Date
    # H = Click through URL
    ad_col = _find_header_column(
        sheet,
        "AD Name",
        "Ad Name",
    )

    notes_col = _find_header_column(
        sheet,
        "Trafficking Notes",
    )

    creative_col = _find_header_column(
        sheet,
        "Creative File Name",
    )

    studio_col = _find_header_column(
        sheet,
        "Studio Creative? (Y/N)",
        "Studio Creative (Y/N)",
    )

    rotation_col = _find_header_column(
        sheet,
        "Rotation %",
        'Rotation % (or "Even")',
        "Rotation",
    )

    start_col = _find_header_column(
        sheet,
        "Start Date",
    )

    end_col = _find_header_column(
        sheet,
        "End Date",
    )

    # Click-through header has extra wording in some versions.
    url_col = None
    for column in range(1, sheet.max_column + 1):
        header = _normalize(
            sheet.cell(
                row=1,
                column=column,
            ).value
        )

        if header.startswith("clickthroughurl"):
            url_col = column
            break

    if url_col is None:
        raise ValueError(
            "Could not find Click through URL column in Multi tab."
        )

    style_snapshot = _snapshot_row_format(
        sheet,
        first_data_row,
        max_column,
    )

    for merged_range in list(
        sheet.merged_cells.ranges
    ):
        if merged_range.min_row >= first_data_row:
            sheet.unmerge_cells(
                str(merged_range)
            )

    old_max_row = sheet.max_row

    _clear_values(
        sheet,
        min_row=first_data_row,
        max_row=old_max_row,
        min_col=1,
        max_col=max_column,
    )

    warnings = []
    output_row = first_data_row

    # One creative block per unique Ad Name.
    ad_blocks: dict[str, dict] = {}

    for record in records:
        try:
            ad_name = build_anthem_ad_name(
                record,
                campaign_name,
            )
        except ValueError as exc:
            warnings.append(str(exc))
            continue

        matches = match_anthem_creatives(
            record,
            creative_names,
        )

        # Multi tab is used only where an Ad has 2+ matching creatives.
        if len(matches) < 2:
            continue

        if ad_name in ad_blocks:
            continue

        start_date, end_date = _resolve_dates(
            record,
            override_start_date,
            override_end_date,
        )

        final_url = _resolve_url(
            record,
            url_map,
        )

        ad_blocks[ad_name] = {
            "matches": matches,
            "start_date": start_date,
            "end_date": end_date,
            "url": final_url,
        }

    for ad_name, block in ad_blocks.items():
        for creative_name in block["matches"]:
            _apply_row_format(
                sheet,
                output_row,
                style_snapshot,
            )

            sheet.cell(
                row=output_row,
                column=ad_col,
            ).value = ad_name

            sheet.cell(
                row=output_row,
                column=notes_col,
            ).value = ""

            # This now writes under the ACTUAL "Creative File Name" header.
            # In the current master_template.xlsm this is column C.
            sheet.cell(
                row=output_row,
                column=creative_col,
            ).value = creative_name

            sheet.cell(
                row=output_row,
                column=studio_col,
            ).value = "N"

            sheet.cell(
                row=output_row,
                column=rotation_col,
            ).value = "Even"

            start_cell = sheet.cell(
                row=output_row,
                column=start_col,
            )
            start_cell.value = block["start_date"]
            start_cell.number_format = "dd-mm-yyyy"

            end_cell = sheet.cell(
                row=output_row,
                column=end_col,
            )
            end_cell.value = block["end_date"]
            end_cell.number_format = "dd-mm-yyyy"

            sheet.cell(
                row=output_row,
                column=url_col,
            ).value = block["url"]

            output_row += 1

        if not block["url"]:
            warnings.append(
                f"URL missing for Multi Ad: {ad_name}"
            )

    return warnings

def _populate_additional_pixels(
    workbook,
) -> None:
    if "Additional Pixels" not in workbook.sheetnames:
        return

    sheet = workbook[
        "Additional Pixels"
    ]

    # Matches the supplied Anthem workbook.
    sheet.cell(
        row=2,
        column=1,
    ).value = "Y"

    sheet.cell(
        row=2,
        column=3,
    ).value = "Double Verify"

    sheet.cell(
        row=2,
        column=4,
    ).value = "Impression & Blocking"

    sheet.cell(
        row=2,
        column=6,
    ).value = (
        "Assembly Control - DV universal pixel"
    )


def generate_anthem_tsheet(
    prisma_file,
    creative_files,
    url_mapping_text: str = "",
    url_mapping: dict | None = None,
    override_start_date=None,
    override_end_date=None,
) -> tuple[bytes, list[str]]:
    """
    Generates an Anthem / Elevance traffic sheet.

    Creative routing rule:
      0 matches -> warning
      1 match   -> creative filename directly in Traffic_Doc
      2+ matches -> Traffic_Doc says SEE MULTI-CREATIVE TAB and
                    every matching creative is added to Multi tab.

    Display matching:
      Language + exact dimension

    Video/CTV matching:
      Language + duration + 16x9

    URL mapping:
      one URL for each Language + Channel combination.

    Example url_mapping_text:
      EN\tDisplay\thttps://...
      SP\tDisplay\thttps://...
      EN\tCTV\thttps://...
      SP\tCTV\thttps://...
      EN\tVideo\thttps://...
      SP\tVideo\thttps://...
    """
    if not MASTER_TEMPLATE.exists():
        raise FileNotFoundError(
            f"master_template.xlsm was not found at: "
            f"{MASTER_TEMPLATE}"
        )

    raw_rows, records = read_prisma_export(
        prisma_file
    )

    campaign_name = _campaign_name_from_prisma(
        raw_rows
    )

    creative_names = _creative_names_from_uploads(
        creative_files
    )

    if not creative_names:
        raise ValueError(
            "No creative files were found in the uploaded files/ZIPs."
        )

    url_map = parse_url_mapping(
        url_mapping_text=url_mapping_text,
        url_mapping=url_mapping,
    )

    # Do not silently generate a sheet when pasted Anthem URLs were not
    # understood. This catches deployment/input issues immediately.
    if url_mapping_text and not url_map:
        raise ValueError(
            "Anthem URLs were pasted, but none could be mapped. "
            "Expected CMP markers DIS, CTV, OLV or DRAD."
        )

    workbook = load_workbook(
        MASTER_TEMPLATE,
        keep_vba=True,
        data_only=False,
    )

    # Paste current Prisma export.
    prisma_sheet = workbook[
        PRISMA_SHEET
    ]

    _clear_values(
        prisma_sheet,
        min_row=1,
        max_row=prisma_sheet.max_row,
        min_col=1,
        max_col=prisma_sheet.max_column,
    )

    _paste_prisma_rows(
        prisma_sheet,
        raw_rows,
    )

    warnings = []

    warnings.extend(
        _populate_traffic_sheet(
            workbook=workbook,
            records=records,
            creative_names=creative_names,
            campaign_name=campaign_name,
            url_map=url_map,
            override_start_date=override_start_date,
            override_end_date=override_end_date,
        )
    )

    warnings.extend(
        _populate_multi_sheet(
            workbook=workbook,
            records=records,
            creative_names=creative_names,
            campaign_name=campaign_name,
            url_map=url_map,
            override_start_date=override_start_date,
            override_end_date=override_end_date,
        )
    )

    _populate_additional_pixels(
        workbook
    )

    output = io.BytesIO()

    workbook.save(output)
    output.seek(0)

    # Remove duplicate warnings while preserving order.
    unique_warnings = list(
        dict.fromkeys(warnings)
    )

    return output.getvalue(), unique_warnings
