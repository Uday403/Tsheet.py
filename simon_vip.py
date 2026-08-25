from __future__ import annotations

import csv
import io
import re
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable


from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
MASTER_TEMPLATE = BASE_DIR / "master_template.xlsm"

PRISMA_SHEET = "Prisma Export - Paste as values"
TRAFFIC_SHEET = "Traffic_Doc"
MULTI_SHEET = "Multi-Ad or Creative Rotation"

TRACKING_1X1 = "Tracking_1x1"


# ============================================================
# BASIC HELPERS
# ============================================================

def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value) -> str:
    """
    Used for matching outlet names / placement text.

    Example:
      "Arundel Mills" -> "arundelmills"
      "ARUNDEL_MILLS" -> "arundelmills"
    """
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _words(value) -> list[str]:
    return [
        word
        for word in re.findall(r"[A-Za-z0-9]+", _clean(value).lower())
        if word
    ]


def _read_uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue

    return data.decode("utf-8", errors="replace")


# ============================================================
# PRISMA CSV
# ============================================================

def _read_csv_rows(uploaded_file) -> list[list[str]]:
    text = _decode_text(_read_uploaded_bytes(uploaded_file))

    try:
        dialect = csv.Sniffer().sniff(
            text[:10000],
            delimiters=",;\t|",
        )
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(text))

    return [list(row) for row in reader]


def _find_prisma_header(
    raw_rows: list[list[str]],
) -> tuple[int, list[str]]:
    """
    Looks for Placement Name without depending on a fixed Prisma column.
    """
    for index, row in enumerate(raw_rows):
        headers = [_clean(cell).replace("\n", " ") for cell in row]
        normalized = {_normalize(header) for header in headers if header}

        if "placementname" in normalized:
            return index, headers

    raise ValueError(
        "The Prisma header row could not be found. "
        "The uploaded file must contain a Placement Name column."
    )


def _record_value(
    record: dict[str, str],
    *possible_headers: str,
) -> str:
    normalized_record = {
        _normalize(key): value
        for key, value in record.items()
    }

    for header in possible_headers:
        value = normalized_record.get(_normalize(header))

        if value is not None and _clean(value):
            return _clean(value)

    return ""


def read_prisma_export(
    uploaded_file,
) -> tuple[list[list[str]], list[dict[str, str]]]:
    """
    Returns:
      raw_rows -> pasted exactly into the Prisma tab
      records  -> only placement rows used for Traffic_Doc
    """
    raw_rows = _read_csv_rows(uploaded_file)
    header_index, headers = _find_prisma_header(raw_rows)

    records: list[dict[str, str]] = []

    for source_index in range(header_index + 1, len(raw_rows)):
        source_row = raw_rows[source_index]

        padded = source_row + [""] * max(
            0,
            len(headers) - len(source_row),
        )

        record = dict(
            zip(
                headers,
                padded[:len(headers)],
            )
        )

        placement_name = _record_value(
            record,
            "Placement Name",
            "Placement name",
        )

        if not placement_name:
            continue

        row_type = _record_value(
            record,
            "Row Type",
            "Type",
            "Package / Placement",
        ).lower()

        if row_type == "package":
            continue

        if placement_name.lower().startswith("package:"):
            continue

        record["_placement_name"] = placement_name
        record["_source_excel_row"] = str(source_index + 1)

        record["_site_name"] = _record_value(
            record,
            "Media outlet / Supplier name (ad server)",
            "Media outlet / Supplier name (Prisma)",
            "Site Name",
            "Site name",
            "Supplier Name",
            "Vendor",
        )

        record["_placement_id"] = _record_value(
            record,
            "Ad server ID",
            "Placement ID",
            "DCM Placement ID",
            "Ad Server Placement ID",
        )

        record["_dimension"] = _record_value(
            record,
            "Dimensions",
            "Dimension",
            "Creative Size",
            "Creative size",
            "Size",
        )

        record["_start_date"] = _record_value(
            record,
            "Flight start date",
            "Start Date",
            "Placement Start Date",
        )

        record["_end_date"] = _record_value(
            record,
            "Flight end date",
            "End Date",
            "Placement End Date",
        )

        records.append(record)

    if not records:
        raise ValueError(
            "Placement Name was found, but no placement rows were detected."
        )

    return raw_rows, records


# ============================================================
# OUTLET + UTM MAPPING
# ============================================================

def parse_outlet_utm_mapping(
    outlet_utm_text: str,
) -> list[dict[str, str]]:
    """
    Dashboard input can be pasted directly from two Excel columns:

        Arundel Mills    https://...
        Desert Hills     https://...
        Jersey Shore     https://...

    Tab-separated is preferred. Comma and pipe are also accepted.
    Header rows such as "Outlet Name | UTM" are ignored.
    """
    mappings: list[dict[str, str]] = []

    for raw_line in outlet_utm_text.splitlines():
        line = raw_line.strip()

        if not line:
            continue

        cells = []

        # Excel copy/paste will normally be tab-separated.
        if "\t" in line:
            cells = line.split("\t")

        elif "|" in line:
            cells = line.split("|", 1)

        elif "," in line:
            # Only split once so commas inside URLs are not damaged.
            cells = line.split(",", 1)

        else:
            # Last fallback: find the first URL in the line.
            match = re.search(r"https?://\S+", line)

            if match:
                cells = [
                    line[:match.start()].strip(),
                    match.group(0).strip(),
                ]

        if len(cells) < 2:
            continue

        outlet = _clean(cells[0])
        url = _clean(cells[1])

        if not outlet or not url:
            continue

        if (
            _normalize(outlet)
            in {
                "outlet",
                "outletname",
                "property",
                "propertyname",
                "mall",
                "mallname",
            }
            and "url" in _normalize(url)
        ):
            continue

        mappings.append(
            {
                "outlet": outlet,
                "outlet_normalized": _normalize(outlet),
                "url": url,
            }
        )

    if not mappings:
        raise ValueError(
            "No Outlet / UTM mappings were detected. "
            "Paste two columns from Excel: Outlet Name and UTM."
        )

    # Longest outlet names first prevents a short name from stealing a match.
    mappings.sort(
        key=lambda item: len(item["outlet_normalized"]),
        reverse=True,
    )

    return mappings


def match_outlet_utm(
    placement_name: str,
    mappings: list[dict[str, str]],
) -> tuple[str, str]:
    """
    Finds the outlet name anywhere inside the Placement Name.

    Example:
      Mapping outlet = Arundel Mills

      Placement =
      Simon_Premium Outlet_USA_Arundel Mills_Q3 Tourism_...

      -> matched.
    """
    normalized_placement = _normalize(placement_name)

    matches = [
        mapping
        for mapping in mappings
        if (
            mapping["outlet_normalized"]
            and mapping["outlet_normalized"] in normalized_placement
        )
    ]

    if not matches:
        return "", ""

    # mappings are already sorted longest-first.
    best = matches[0]

    return best["outlet"], best["url"]


# ============================================================
# DIMENSION + CREATIVE MATCHING
# ============================================================

def _extract_dimension(value: str) -> str:
    match = re.search(
        r"(?<!\d)(\d{1,4})\s*[xX×*]\s*(\d{1,4})(?!\d)",
        _clean(value),
    )

    if not match:
        return ""

    return f"{match.group(1)}x{match.group(2)}"


def placement_dimension(record: dict[str, str]) -> str:
    dimension = _extract_dimension(
        record.get("_dimension", "")
    )

    if dimension:
        return dimension

    return _extract_dimension(
        record.get("_placement_name", "")
    )


def _creative_file_names(
    creative_files: Iterable | None,
) -> list[str]:
    """
    Supports individual creatives and ZIP files.
    """
    names: list[str] = []

    for uploaded_file in creative_files or []:
        file_name = Path(uploaded_file.name).name

        if file_name.lower().endswith(".zip"):
            uploaded_file.seek(0)

            with zipfile.ZipFile(
                uploaded_file,
                "r",
            ) as archive:
                for member in archive.namelist():
                    if member.endswith("/"):
                        continue

                    name = Path(member).name

                    if name:
                        names.append(name)

        else:
            names.append(file_name)

    return list(
        dict.fromkeys(
            name
            for name in names
            if name
        )
    )


def _creative_content_tokens(
    creative_name: str,
) -> set[str]:
    """
    Removes dimension and generic file words, leaving useful content words.
    """
    stem = Path(creative_name).stem

    stem = re.sub(
        r"(?<!\d)\d{1,4}\s*[xX×*]\s*\d{1,4}(?!\d)",
        " ",
        stem,
    )

    generic = {
        "simon",
        "premium",
        "outlet",
        "outlets",
        "creative",
        "display",
        "banner",
        "static",
        "image",
        "jpg",
        "jpeg",
        "png",
        "gif",
        "webp",
        "ad",
        "ads",
        "version",
        "ver",
        "new",
    }

    return {
        word
        for word in _words(stem)
        if len(word) >= 3 and word not in generic
    }


def _creative_score(
    creative_name: str,
    placement_name: str,
    dimension: str,
) -> float:
    creative_dimension = _extract_dimension(
        creative_name
    )

    # When placement has a real display dimension,
    # the creative must have the same dimension.
    if dimension and dimension.lower() != "1x1":
        if not creative_dimension:
            return -1000

        if creative_dimension.lower() != dimension.lower():
            return -1000

    score = 0.0

    if (
        dimension
        and creative_dimension
        and creative_dimension.lower() == dimension.lower()
    ):
        score += 50

    normalized_placement = _normalize(
        placement_name
    )

    # Strong content/name matching.
    for token in _creative_content_tokens(
        creative_name
    ):
        if _normalize(token) in normalized_placement:
            score += 10

    # Longer chunks get stronger weight.
    stem = Path(creative_name).stem

    for chunk in re.split(
        r"[_\-\s]+",
        stem,
    ):
        normalized_chunk = _normalize(chunk)

        if (
            len(normalized_chunk) >= 6
            and normalized_chunk in normalized_placement
        ):
            score += 15

    return score


def match_creative(
    creative_names: list[str],
    placement_name: str,
    dimension: str,
) -> tuple[str, list[str]]:
    """
    Returns:
      creative name
      tied candidates (if ambiguous)

    IMPORTANT:
      If NO creative files were uploaded, caller uses Tracking_1x1.
    """
    if not creative_names:
        return TRACKING_1X1, []

    ranked = sorted(
        (
            (
                _creative_score(
                    creative_name=name,
                    placement_name=placement_name,
                    dimension=dimension,
                ),
                name,
            )
            for name in creative_names
        ),
        reverse=True,
    )

    valid = [
        item
        for item in ranked
        if item[0] > 0
    ]

    if not valid:
        # If this is genuinely a 1x1 placement, Tracking_1x1 is the safe fallback.
        if dimension.lower() == "1x1":
            return TRACKING_1X1, []

        return "", []

    best_score = valid[0][0]

    tied = [
        name
        for score, name in valid
        if score == best_score
    ]

    if len(tied) > 1:
        return "", tied

    return valid[0][1], []


# ============================================================
# TEMPLATE / EXCEL HELPERS
# ============================================================

def _find_traffic_layout(
    sheet,
) -> tuple[int, int]:
    """
    Simon sample:
      header row = 8
      first data row = 9

    We find it dynamically so the script does not depend on row 8 forever.
    """
    for row in range(
        1,
        min(sheet.max_row, 30) + 1,
    ):
        placement_header = _clean(
            sheet.cell(row=row, column=4).value
        ).lower()

        ad_header = _clean(
            sheet.cell(row=row, column=8).value
        ).lower()

        if (
            "placement name" in placement_header
            and "ad name" in ad_header
        ):
            return row, row + 1

    raise ValueError(
        "Traffic_Doc headers could not be located."
    )


def _snapshot_row_format(
    sheet,
    row_number: int,
    max_column: int,
) -> dict:
    snapshot = {}

    for column in range(
        1,
        max_column + 1,
    ):
        cell = sheet.cell(
            row=row_number,
            column=column,
        )

        snapshot[column] = {
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }

    return {
        "cells": snapshot,
        "height": sheet.row_dimensions[
            row_number
        ].height,
    }


def _apply_row_format(
    sheet,
    row_number: int,
    snapshot: dict,
) -> None:
    for column, style in snapshot[
        "cells"
    ].items():
        cell = sheet.cell(
            row=row_number,
            column=column,
        )

        cell._style = copy(
            style["style"]
        )
        cell.number_format = style[
            "number_format"
        ]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(
            style["border"]
        )
        cell.alignment = copy(
            style["alignment"]
        )
        cell.protection = copy(
            style["protection"]
        )

    sheet.row_dimensions[
        row_number
    ].height = snapshot["height"]


def _clear_values(
    sheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def _clear_prisma_sheet(
    sheet,
) -> None:
    _clear_values(
        sheet,
        min_row=1,
        max_row=max(
            sheet.max_row,
            200,
        ),
        min_col=1,
        max_col=max(
            sheet.max_column,
            60,
        ),
    )


def _paste_prisma_rows(
    sheet,
    raw_rows: list[list[str]],
) -> None:
    for row_number, row_values in enumerate(
        raw_rows,
        start=1,
    ):
        for column_number, value in enumerate(
            row_values,
            start=1,
        ):
            sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )


def _clear_multi_sheet(
    workbook,
) -> None:
    if MULTI_SHEET not in workbook.sheetnames:
        return

    sheet = workbook[MULTI_SHEET]

    # Unmerge old rotation blocks first.
    for merged_range in list(
        sheet.merged_cells.ranges
    ):
        if merged_range.min_row >= 2:
            sheet.unmerge_cells(
                str(merged_range)
            )

    _clear_values(
        sheet,
        min_row=2,
        max_row=max(
            sheet.max_row,
            100,
        ),
        min_col=1,
        max_col=max(
            sheet.max_column,
            16,
        ),
    )


def _to_excel_date(
    value,
):
    if value is None:
        return ""

    if isinstance(
        value,
        (
            datetime,
        ),
    ):
        return value

    text = _clean(value)

    if not text:
        return ""

    for fmt in (
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(
                text,
                fmt,
            )
        except ValueError:
            pass

    return text


def _campaign_name_from_prisma(
    raw_rows: list[list[str]],
) -> str:
    for row in raw_rows:
        if not row:
            continue

        first = _clean(
            row[0]
        ).lower()

        if first in {
            "campaign name",
            "campaign name:",
        }:
            return (
                _clean(row[1])
                if len(row) > 1
                else ""
            )

    return ""


# ============================================================
# SIMON VIP PREVIEW
# ============================================================

def preview_simon_vip_setup(
    prisma_file,
    creative_files,
    outlet_utm_text: str,
) -> dict:
    _, records = read_prisma_export(
        prisma_file
    )

    mappings = parse_outlet_utm_mapping(
        outlet_utm_text
    )

    creative_names = _creative_file_names(
        creative_files
    )

    rows = []
    warnings = []

    for record in records:
        placement_name = record[
            "_placement_name"
        ]

        dimension = placement_dimension(
            record
        )

        outlet, url = match_outlet_utm(
            placement_name,
            mappings,
        )

        creative, tied = match_creative(
            creative_names=creative_names,
            placement_name=placement_name,
            dimension=dimension,
        )

        if not outlet:
            warnings.append(
                f"No outlet/UTM match: {placement_name}"
            )

        if tied:
            warnings.append(
                f"Ambiguous creative match: "
                f"{placement_name} -> "
                + ", ".join(tied)
            )

        if not creative:
            warnings.append(
                f"No creative matched: {placement_name}"
            )

        rows.append(
            {
                "placement_name": placement_name,
                "ad_name": placement_name,
                "dimension": dimension,
                "outlet": outlet,
                "url": url,
                "creative": creative,
                "creative_candidates": tied,
            }
        )

    matched_utm_count = sum(
        1
        for row in rows
        if row["url"]
    )

    return {
        "rows": rows,
        "outlet_mapping_count": len(
            mappings
        ),
        "placement_count": len(
            records
        ),
        "utm_matched_count": matched_utm_count,
        "utm_unmatched_count": (
            len(records)
            - matched_utm_count
        ),
        "creative_count": len(
            creative_names
        ),
        "using_tracking_1x1": (
            len(creative_names) == 0
        ),
        "warnings": warnings,
    }


# ============================================================
# GENERATE TRAFFIC DOC
# ============================================================

def _populate_traffic_doc(
    workbook,
    records: list[dict[str, str]],
    creative_names: list[str],
    mappings: list[dict[str, str]],
) -> list[str]:
    sheet = workbook[
        TRAFFIC_SHEET
    ]

    _, first_data_row = (
        _find_traffic_layout(sheet)
    )

    max_column = max(
        sheet.max_column,
        24,
    )

    # Preserve style BEFORE clearing old data.
    style_snapshot = (
        _snapshot_row_format(
            sheet,
            first_data_row,
            max_column,
        )
    )

    # Completely remove old campaign values.
    _clear_values(
        sheet,
        min_row=first_data_row,
        max_row=max(
            sheet.max_row,
            first_data_row + len(records) + 10,
        ),
        min_col=1,
        max_col=max_column,
    )

    warnings: list[str] = []

    for index, record in enumerate(
        records
    ):
        row = first_data_row + index

        _apply_row_format(
            sheet,
            row,
            style_snapshot,
        )

        placement_name = record[
            "_placement_name"
        ]

        dimension = placement_dimension(
            record
        )

        outlet_name, matched_url = (
            match_outlet_utm(
                placement_name,
                mappings,
            )
        )

        creative_name, tied = (
            match_creative(
                creative_names=creative_names,
                placement_name=placement_name,
                dimension=dimension,
            )
        )

        # Simon VIP confirmed rule:
        # Placement Name = Ad Name.
        ad_name = placement_name

        if not matched_url:
            warnings.append(
                f"No UTM matched outlet in placement: "
                f"{placement_name}"
            )

        if tied:
            warnings.append(
                f"Ambiguous creative match — "
                f"manual review required: "
                f"{placement_name} -> "
                + ", ".join(tied)
            )

        if not creative_name:
            warnings.append(
                f"No creative matched: "
                f"{placement_name}"
            )

        # ----------------------------------------------------
        # Traffic_Doc columns based on shared template:
        # B Site Name
        # C DCM Placement ID
        # D Placement Name
        # E Dimensions
        # H Ad Name
        # J Action
        # K Creative
        # L Studio Creative
        # M Rotation
        # N Start
        # O End
        # P URL
        # ----------------------------------------------------

        sheet.cell(
            row=row,
            column=2,
        ).value = record[
            "_site_name"
        ]

        sheet.cell(
            row=row,
            column=3,
        ).value = record[
            "_placement_id"
        ]

        sheet.cell(
            row=row,
            column=4,
        ).value = placement_name

        sheet.cell(
            row=row,
            column=5,
        ).value = dimension

        sheet.cell(
            row=row,
            column=8,
        ).value = ad_name

        sheet.cell(
            row=row,
            column=10,
        ).value = "New"

        sheet.cell(
            row=row,
            column=11,
        ).value = creative_name

        # Sample Simon sheet uses no Studio value.
        sheet.cell(
            row=row,
            column=12,
        ).value = None

        # Single creative rotation.
        sheet.cell(
            row=row,
            column=13,
        ).value = "100%"

        sheet.cell(
            row=row,
            column=14,
        ).value = _to_excel_date(
            record["_start_date"]
        )

        sheet.cell(
            row=row,
            column=15,
        ).value = _to_excel_date(
            record["_end_date"]
        )

        sheet.cell(
            row=row,
            column=16,
        ).value = matched_url

    return warnings


# ============================================================
# MAIN ENTRY POINT USED BY Tsheet.py
# ============================================================

def generate_simon_vip_tsheet(
    prisma_file,
    creative_files,
    outlet_utm_text: str,
) -> tuple[bytes, list[str]]:
    """
    Simon VIP rules implemented:

    1. Placement Name = Ad Name.

    2. Outlet / UTM mapping is pasted from two Excel columns.
       Outlet name is matched anywhere inside Placement Name.

    3. If creative files are uploaded:
       - match by placement dimension
       - then content/name similarity.

    4. If NO creatives are uploaded:
       - Creative File Name = Tracking_1x1.

    5. UTM is used exactly as provided.
       Simon VIP does NOT generate or modify the UTM.

    6. Old Traffic_Doc / Prisma / Multi-tab campaign data is cleared.
    """
    if not MASTER_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Master template not found: "
            f"{MASTER_TEMPLATE.name}"
        )

    raw_rows, records = (
        read_prisma_export(
            prisma_file
        )
    )

    mappings = (
        parse_outlet_utm_mapping(
            outlet_utm_text
        )
    )

    creative_names = (
        _creative_file_names(
            creative_files
        )
    )

    workbook = load_workbook(
        MASTER_TEMPLATE,
        keep_vba=True,
    )

    for required_sheet in (
        PRISMA_SHEET,
        TRAFFIC_SHEET,
    ):
        if (
            required_sheet
            not in workbook.sheetnames
        ):
            raise KeyError(
                f"Missing worksheet in "
                f"master template: "
                f"{required_sheet}"
            )

    # Clear + paste Prisma.
    prisma_sheet = workbook[
        PRISMA_SHEET
    ]

    _clear_prisma_sheet(
        prisma_sheet
    )

    _paste_prisma_rows(
        prisma_sheet,
        raw_rows,
    )

    # Remove stale Multi-tab data.
    _clear_multi_sheet(
        workbook
    )

    # Campaign name if Prisma export provides one.
    campaign_name = (
        _campaign_name_from_prisma(
            raw_rows
        )
    )

    if campaign_name:
        workbook[
            TRAFFIC_SHEET
        ]["B1"] = campaign_name

    warnings = (
        _populate_traffic_doc(
            workbook=workbook,
            records=records,
            creative_names=creative_names,
            mappings=mappings,
        )
    )

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output = io.BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return (
        output.getvalue(),
        warnings,
    )
