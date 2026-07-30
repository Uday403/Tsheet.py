from __future__ import annotations

import csv
import io
import re
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
MASTER_TEMPLATE = BASE_DIR / "master_template.xlsm"

PRISMA_SHEET = "Prisma Export - Paste as values"
TRAFFIC_SHEET = "Traffic_Doc"
ROTATION_SHEET = "Multi-Ad or Creative Rotation"

TRAFFIC_HEADER_ROW = 6
TRAFFIC_FIRST_DATA_ROW = 7
TRAFFIC_LAST_COLUMN = 23  # Column W


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _normalize(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _read_uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_csv_rows(uploaded_file) -> list[list[str]]:
    text = _decode_text(_read_uploaded_bytes(uploaded_file))

    try:
        dialect = csv.Sniffer().sniff(
            text[:10000],
            delimiters=",;\t|",
        )
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(text))

    return [list(row) for row in reader]


def _find_header_row(raw_rows: list[list[str]]) -> tuple[int, list[str]]:
    for index, row in enumerate(raw_rows):
        headers = [_clean(cell).replace("\n", " ") for cell in row]
        normalized = {_normalize(header) for header in headers}

        if "placementname" in normalized:
            return index, headers

    raise ValueError(
        "The Prisma header row could not be found. "
        "The uploaded file must contain a Placement Name column."
    )


def _header_value(
    record: dict[str, str],
    *possible_names: str,
) -> str:
    normalized_record = {
        _normalize(key): value
        for key, value in record.items()
    }

    for name in possible_names:
        value = normalized_record.get(_normalize(name))
        if value is not None:
            return _clean(value)

    return ""


def read_prisma_export(
    uploaded_file,
) -> tuple[list[list[str]], list[dict[str, str]]]:
    """
    Returns:
      1. Every row from the Prisma export, for pasting into the Prisma tab.
      2. Only valid placement records, including the original Excel row number.

    Package/header rows are not sent to Traffic_Doc.
    """
    raw_rows = _read_csv_rows(uploaded_file)
    header_index, headers = _find_header_row(raw_rows)

    records: list[dict[str, str]] = []

    for raw_index in range(header_index + 1, len(raw_rows)):
        row = raw_rows[raw_index]
        padded = row + [""] * max(0, len(headers) - len(row))
        record = dict(zip(headers, padded[:len(headers)]))

        placement_name = _header_value(record, "Placement Name")
        if not placement_name:
            continue

        # Ignore obvious package rows.
        row_type = _header_value(
            record,
            "Row Type",
            "Type",
            "Package / Placement",
        ).lower()

        if row_type == "package":
            continue

        normalized_name = placement_name.lower()
        if normalized_name.startswith("package:"):
            continue

        record["Placement Name"] = placement_name

        # Excel row number after pasting the full raw export at row 1.
        record["_source_excel_row"] = str(raw_index + 1)
        records.append(record)

    if not records:
        raise ValueError(
            "The Placement Name header was found, but no placement rows "
            "were detected below it."
        )

    return raw_rows, records


def _clear_sheet_values(
    sheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def paste_prisma_export(
    workbook,
    raw_rows: list[list[str]],
) -> None:
    if PRISMA_SHEET not in workbook.sheetnames:
        raise KeyError(f"Missing worksheet: {PRISMA_SHEET}")

    sheet = workbook[PRISMA_SHEET]

    _clear_sheet_values(
        sheet,
        min_row=1,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max(sheet.max_column, 60),
    )

    for row_number, row_values in enumerate(raw_rows, start=1):
        for column_number, value in enumerate(row_values, start=1):
            sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )


def _find_dimension(text: str) -> str:
    match = re.search(
        r"(?<!\d)(\d{2,4})\s*[xX]\s*(\d{2,4})(?!\d)",
        _clean(text),
    )
    return f"{match.group(1)}x{match.group(2)}" if match else ""


def _placement_parts(placement_name: str) -> list[str]:
    return [_clean(part) for part in placement_name.split("_")]


def _parse_pulte_placement(placement_name: str) -> dict[str, str]:
    parts = _placement_parts(placement_name)

    brand_lookup = {
        "pulte": "Pulte",
        "delwebb": "Del Webb",
        "centex": "Centex",
        "divosta": "DiVosta",
        "johnwieland": "John Wieland",
        "wieland": "Wieland",
    }

    brand_index = None
    brand = "Pulte"

    for index, part in enumerate(parts):
        normalized_part = _normalize(part)
        if normalized_part in brand_lookup:
            brand_index = index
            brand = brand_lookup[normalized_part]
            break

    division = ""
    campaign = ""
    community = ""

    if brand_index is not None:
        if brand_index >= 1:
            division = parts[brand_index - 1]
        if brand_index + 1 < len(parts):
            campaign = parts[brand_index + 1]
        if brand_index + 2 < len(parts):
            community = parts[brand_index + 2]

    community_id = ""
    for part in reversed(parts):
        match = re.search(r"(?<!\d)(\d{5,7})(?!\d)", part)
        if match:
            community_id = match.group(1)
            break

    return {
        "division": division,
        "brand": brand,
        "campaign": campaign,
        "community": community,
        "community_id": community_id,
        "dimension": _find_dimension(placement_name),
    }


def build_ad_name(placement_name: str) -> str:
    parsed = _parse_pulte_placement(placement_name)

    values = [
        parsed["division"],
        parsed["brand"],
        parsed["campaign"],
        parsed["community"],
        parsed["dimension"],
    ]

    return "_".join(value for value in values if value)


def _creative_file_names(creative_files: Iterable) -> list[str]:
    return [
        Path(file.name).name
        for file in creative_files
        if getattr(file, "name", None)
    ]


def _creative_score(
    creative_name: str,
    placement_name: str,
) -> int:
    parsed = _parse_pulte_placement(placement_name)
    normalized_creative = _normalize(creative_name)
    score = 0

    weighted_values = (
        (parsed["community_id"], 30),
        (parsed["community"], 22),
        (parsed["dimension"], 20),
        (parsed["division"], 6),
        (parsed["brand"], 4),
    )

    for value, weight in weighted_values:
        normalized_value = _normalize(value)
        if normalized_value and normalized_value in normalized_creative:
            score += weight

    for word in re.split(r"\W+", parsed["community"].lower()):
        if len(word) >= 4 and word in creative_name.lower():
            score += 4

    return score


def match_creative(
    creative_names: list[str],
    placement_name: str,
) -> str:
    if not creative_names:
        return ""

    ranked = sorted(
        (
            (_creative_score(name, placement_name), name)
            for name in creative_names
        ),
        key=lambda item: (item[0], item[1]),
        reverse=True,
    )

    return ranked[0][1] if ranked[0][0] > 0 else ""


def _parse_complete_urls(text: str) -> list[str]:
    """
    Accepts:
      - one complete URL per line
      - placement<TAB>URL
      - ad name<TAB>URL
      - community ID<TAB>URL

    The URL is never modified.
    """
    urls: list[str] = []

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line:
            urls.append(line)

    return urls


def _extract_url(line: str) -> str:
    match = re.search(r"https?://\S+", line)
    return match.group(0).rstrip(",;") if match else ""


def match_complete_url(
    url_lines: list[str],
    placement_name: str,
    ad_name: str,
    placement_index: int,
    total_placements: int,
) -> str:
    parsed = _parse_pulte_placement(placement_name)

    community_id = parsed["community_id"]
    normalized_placement = _normalize(placement_name)
    normalized_ad_name = _normalize(ad_name)

    # First: mapped line containing placement/ad/community ID.
    for line in url_lines:
        url = _extract_url(line)
        if not url:
            continue

        normalized_line = _normalize(line)

        if community_id and community_id in line:
            return url
        if normalized_ad_name and normalized_ad_name in normalized_line:
            return url
        if normalized_placement and normalized_placement in normalized_line:
            return url

    # Second: community words embedded in URL.
    community_words = [
        word.lower()
        for word in re.split(r"\W+", parsed["community"])
        if len(word) >= 4
    ]

    for line in url_lines:
        url = _extract_url(line)
        lowered_url = url.lower()

        if url and community_words and all(
            word in lowered_url
            for word in community_words
        ):
            return url

    # Third: same number of URLs and placements = map by order.
    extracted_urls = [
        _extract_url(line)
        for line in url_lines
        if _extract_url(line)
    ]

    if len(extracted_urls) == total_placements:
        return extracted_urls[placement_index]

    # Fourth: only one URL supplied = apply to all.
    if len(extracted_urls) == 1:
        return extracted_urls[0]

    return ""


def _copy_row_style(
    sheet,
    source_row: int,
    target_row: int,
) -> None:
    for column in range(1, TRAFFIC_LAST_COLUMN + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)

        if source.has_style:
            target._style = copy(source._style)

        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)

    sheet.row_dimensions[target_row].height = (
        sheet.row_dimensions[source_row].height
    )


def _clear_old_traffic_rows(sheet) -> None:
    """
    Clears every old Traffic_Doc value below the header.
    Styles remain available in the workbook.
    """
    _clear_sheet_values(
        sheet,
        min_row=TRAFFIC_FIRST_DATA_ROW,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max(sheet.max_column, TRAFFIC_LAST_COLUMN),
    )


def _clear_rotation_sheet(workbook) -> None:
    if ROTATION_SHEET not in workbook.sheetnames:
        return

    sheet = workbook[ROTATION_SHEET]

    _clear_sheet_values(
        sheet,
        min_row=2,
        max_row=max(sheet.max_row, 5000),
        min_col=1,
        max_col=max(sheet.max_column, 10),
    )


def _traffic_formulas(
    traffic_row: int,
    prisma_row: int,
) -> dict[int, str]:
    """
    Builds the formulas used by the shared master template.

    The Prisma row is the original placement row from the pasted export,
    so package rows do not break the Traffic_Doc references.
    """
    r = traffic_row
    p = prisma_row

    return {
        1: (
            f'=IF(D{r}<>"",CONCATENATE('
            f'IF(OR(\'Additional Pixels\'!A$2="Y",'
            f'\'Additional Pixels\'!A$2="Yes"),'
            f'\'Additional Pixels\'!B$2,""),'
            f'IF(OR(\'Additional Pixels\'!A$3="Y",'
            f'\'Additional Pixels\'!A$3="Yes"),'
            f'", "&\'Additional Pixels\'!B$3,""),'
            f'IF(OR(\'Additional Pixels\'!A$4="Y",'
            f'\'Additional Pixels\'!A$4="Yes"),'
            f'", "&\'Additional Pixels\'!B$4,""),'
            f'IF(OR(\'Additional Pixels\'!A$5="Y",'
            f'\'Additional Pixels\'!A$5="Yes"),'
            f'", "&\'Additional Pixels\'!B$5,""),'
            f'IF(OR(\'Additional Pixels\'!A$6="Y",'
            f'\'Additional Pixels\'!A$6="Yes"),'
            f'", "&\'Additional Pixels\'!B$6,""),'
            f'IF(OR(\'Additional Pixels\'!A$7="Y",'
            f'\'Additional Pixels\'!A$7="Yes"),'
            f'", "&\'Additional Pixels\'!B$7,""),'
            f'IF(OR(\'Additional Pixels\'!A$8="Y",'
            f'\'Additional Pixels\'!A$8="Yes"),'
            f'", "&\'Additional Pixels\'!B$8,""),'
            f'IF(OR(\'Additional Pixels\'!A$9="Y",'
            f'\'Additional Pixels\'!A$9="Yes"),'
            f'", "&\'Additional Pixels\'!B$9,""),'
            f'IF(OR(\'Additional Pixels\'!A$10="Y",'
            f'\'Additional Pixels\'!A$10="Yes"),'
            f'", "&\'Additional Pixels\'!B$10,""),'
            f'IF(OR(\'Additional Pixels\'!A$11="Y",'
            f'\'Additional Pixels\'!A$11="Yes"),'
            f'", "&\'Additional Pixels\'!B$11,"")),"")'
        ),
        2: (
            f'=IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(\'{PRISMA_SHEET}\'!J{p}="","",'
            f'\'{PRISMA_SHEET}\'!J{p}))'
        ),
        3: (
            f'=IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(\'{PRISMA_SHEET}\'!O{p}="","",'
            f'\'{PRISMA_SHEET}\'!O{p}))'
        ),
        4: (
            f'=IF(AND(\'{PRISMA_SHEET}\'!U{p}="",'
            f'\'{PRISMA_SHEET}\'!W{p}=""),"",'
            f'IF(\'{PRISMA_SHEET}\'!U{p}="",'
            f'"DELETE BLANK COLUMN Q - SEE INSTRUCTION TAB",'
            f'\'{PRISMA_SHEET}\'!U{p}))'
        ),
        5: (
            f'=IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(\'{PRISMA_SHEET}\'!T{p}="","",'
            f'SUBSTITUTE(\'{PRISMA_SHEET}\'!T{p}," ","")))'
        ),
        6: (
            f'=IFERROR(IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(ISNUMBER(MID(D{r},FIND(":",D{r})+1,2))="TRUE",'
            f'MID(D{r},FIND(":",D{r})+1,2),'
            f'IF(ISNUMBER(SEARCH("*15s*",D{r})),"15",'
            f'IF(ISNUMBER(SEARCH("*30s*",D{r})),"30",'
            f'IF(ISNUMBER(SEARCH("*60s*",D{r})),"60",'
            f'IF(ISNUMBER(SEARCH("*90s*",D{r})),"90",'
            f'IF(ISNUMBER(SEARCH("*6s*",D{r})),"6",""))))))),"")'
        ),
        7: (
            f'=IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(E{r}="0x0",'
            f'IF(ISNUMBER(SEARCH("HULU",D{r})),"Vast",'
            f'IF(ISNUMBER(SEARCH("ROKU",D{r})),"Vast",'
            f'IF(ISNUMBER(SEARCH("ESPN",D{r})),"Vast","Vpaid"))),""))'
        ),
        9: (
            f'=IF(ISNUMBER(SEARCH("*DELETE*",D{r})),"",'
            f'IF(K{r}="","",IF(E{r}="1x1","",'
            f'IF(F{r}="",'
            f'IF(COUNTIF(K{r},"*"&E{r}&"*")>0,"","CHECK DIMENSION"),'
            f'IF(COUNTIF(K{r},"*"&F{r}&"*")>0,"",'
            f'IF(AND(B{r}="*NBC*",D{r}="*Hulu*"),'
            f'"Select ONLY 1920x1080 (15,00-30,000 kbps bitrate) on creative",'
            f'"CHECK DIMENSIONS"))))))'
        ),
        13: (
            f'=IF(H{r}="","100%",'
            f'IF(COUNTIF(\'{ROTATION_SHEET}\'!A:A,H{r})>0,'
            f'"SEE MULTI TAB",'
            f'IFERROR(IF(COUNTIF(\'{ROTATION_SHEET}\'!A:A,'
            f'LEFT(H{r},FIND(",",H{r})-1))>0,'
            f'"SEE MULTI TAB","100%"),"100%")))'
        ),
        14: (
            f'=IF(AND(\'{PRISMA_SHEET}\'!W{p}="",'
            f'\'{PRISMA_SHEET}\'!W{p}=""),"",'
            f'IF(\'{PRISMA_SHEET}\'!W{p}="",'
            f'"DELETE BLANK COLUMN Q - SEE INSTRUCTION TAB",'
            f'\'{PRISMA_SHEET}\'!W{p}))'
        ),
        15: (
            f'=IF(AND(\'{PRISMA_SHEET}\'!X{p}="",'
            f'\'{PRISMA_SHEET}\'!X{p}=""),"",'
            f'IF(\'{PRISMA_SHEET}\'!X{p}="",'
            f'"DELETE BLANK COLUMN Q - SEE INSTRUCTION TAB",'
            f'\'{PRISMA_SHEET}\'!X{p}))'
        ),
    }


def populate_pulte_normal(
    workbook,
    records: list[dict[str, str]],
    creative_files: Iterable,
    complete_urls_text: str,
) -> list[str]:
    if TRAFFIC_SHEET not in workbook.sheetnames:
        raise KeyError(f"Missing worksheet: {TRAFFIC_SHEET}")

    sheet = workbook[TRAFFIC_SHEET]

    # Preserve one clean style row before clearing old output.
    style_source_row = TRAFFIC_FIRST_DATA_ROW
    style_snapshot = {}

    for column in range(1, TRAFFIC_LAST_COLUMN + 1):
        cell = sheet.cell(style_source_row, column)
        style_snapshot[column] = {
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }

    source_height = sheet.row_dimensions[style_source_row].height

    _clear_old_traffic_rows(sheet)
    _clear_rotation_sheet(workbook)

    creative_names = _creative_file_names(creative_files)
    url_lines = _parse_complete_urls(complete_urls_text)
    warnings: list[str] = []

    for index, record in enumerate(records):
        traffic_row = TRAFFIC_FIRST_DATA_ROW + index
        prisma_row = int(record["_source_excel_row"])
        placement_name = record["Placement Name"]

        # Restore the template formatting.
        for column in range(1, TRAFFIC_LAST_COLUMN + 1):
            target = sheet.cell(traffic_row, column)
            snapshot = style_snapshot[column]
            target._style = copy(snapshot["style"])
            target.number_format = snapshot["number_format"]
            target.font = copy(snapshot["font"])
            target.fill = copy(snapshot["fill"])
            target.border = copy(snapshot["border"])
            target.alignment = copy(snapshot["alignment"])
            target.protection = copy(snapshot["protection"])

        sheet.row_dimensions[traffic_row].height = source_height

        # Formula-driven columns.
        for column, formula in _traffic_formulas(
            traffic_row=traffic_row,
            prisma_row=prisma_row,
        ).items():
            sheet.cell(traffic_row, column).value = formula

        # Manual columns.
        ad_name = build_ad_name(placement_name)
        creative_name = match_creative(
            creative_names,
            placement_name,
        )
        complete_url = match_complete_url(
            url_lines=url_lines,
            placement_name=placement_name,
            ad_name=ad_name,
            placement_index=index,
            total_placements=len(records),
        )

        sheet.cell(traffic_row, 8).value = ad_name       # H: AD Name
        sheet.cell(traffic_row, 10).value = "New"       # J: Action
        sheet.cell(traffic_row, 11).value = creative_name  # K
        sheet.cell(traffic_row, 12).value = "N"         # L
        sheet.cell(traffic_row, 16).value = complete_url  # P

        if not creative_name:
            warnings.append(
                f"No creative matched: {placement_name}"
            )
        else:
            placement_dimension = _find_dimension(placement_name)
            creative_dimension = _find_dimension(creative_name)

            if (
                placement_dimension
                and creative_dimension
                and placement_dimension.lower()
                != creative_dimension.lower()
            ):
                warnings.append(
                    f"Dimension mismatch — placement "
                    f"{placement_dimension}, creative "
                    f"{creative_dimension}: {placement_name}"
                )

        if not complete_url:
            warnings.append(
                f"No complete URL/UTM matched: {placement_name}"
            )

    return warnings


def generate_normal_pulte_tsheet(
    prisma_file,
    creative_files,
    complete_urls_text: str,
) -> tuple[bytes, list[str]]:
    """
    Entry point used by Tsheet.py.

    Normal Pulte rules:
      - Uses the same master_template.xlsm as every other account.
      - Uses the complete URL/UTM supplied by the team.
      - Never generates, appends, or changes the UTM.
      - Generates Ad Name.
      - Matches Creative File Name.
      - Keeps the template's CHECK DIMENSION formula.
    """
    if not MASTER_TEMPLATE.exists():
        raise FileNotFoundError(
            f"{MASTER_TEMPLATE.name} was not found in the GitHub repository."
        )

    raw_rows, records = read_prisma_export(prisma_file)

    workbook = load_workbook(
        MASTER_TEMPLATE,
        keep_vba=True,
    )

    paste_prisma_export(
        workbook,
        raw_rows,
    )

    warnings = populate_pulte_normal(
        workbook=workbook,
        records=records,
        creative_files=creative_files,
        complete_urls_text=complete_urls_text,
    )

    # Ask Excel to recalculate formulas when the downloaded file is opened.
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue(), warnings
